"""
Atualiza fatores de produto a partir de planilha (ex.: Palete.xlsx):

- conversion_factor — UN por 1 CX (opcional se houver coluna reconhecida)
- pallet_conversion_factor — CX por 1 PL (opcional se houver coluna reconhecida)

É necessário: coluna de código + pelo menos uma coluna de fator (UN/CX e/ou CX/PL).
Colunas são mapeadas pelos mesmos aliases da importação em Cadastro (HEADER_ALIASES).

Uso (na raiz do projeto, com DATABASE_URL no ambiente ou .env):
  python scripts/apply_palete_un_por_cx_xlsx.py "C:\\Users\\...\\Palete.xlsx"
  python scripts/apply_palete_un_por_cx_xlsx.py caminho.xlsx --dry-run
"""
from __future__ import annotations

import argparse
import math
import os
import sys
from pathlib import Path

# Raiz do repositório no sys.path
_ROOT = Path(__file__).resolve().parents[1]
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

os.chdir(_ROOT)

from dotenv import load_dotenv
from openpyxl import load_workbook
from sqlmodel import Session, select

load_dotenv()

from app.api.routes.products import (  # noqa: E402
    HEADER_ALIASES,
    _coerce_conversion_factor_in_row,
    _coerce_pallet_conversion_factor_in_row,
    _ensure_product_pallet_conversion_factor_column,
    _norm_header,
    _normalize_codigo_import,
    _safe_log_change,
)
from app.db.session import SessionLocal  # noqa: E402
from app.models import Product, ProductHistory  # noqa: E402
from app.models.entities import utcnow  # noqa: E402


def _norm_cod_cell(raw) -> str:
    d: dict = {"cod_produto": raw}
    _normalize_codigo_import(d)
    return (d.get("cod_produto") or "").strip()


def _factor_un_per_cx_from_cell(raw) -> float | None:
    d: dict = {"conversion_factor": raw}
    _coerce_conversion_factor_in_row(d)
    v = d.get("conversion_factor")
    return float(v) if v is not None else None


def _factor_cx_per_pl_from_cell(raw) -> float | None:
    d: dict = {"pallet_conversion_factor": raw}
    _coerce_pallet_conversion_factor_in_row(d)
    v = d.get("pallet_conversion_factor")
    return float(v) if v is not None else None


def _code_lookup_variants(code: str) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(x: str) -> None:
        x = (x or "").strip()
        if x and x not in seen:
            seen.add(x)
            out.append(x)

    add(code)
    try:
        n = int(float(code.replace(",", ".")))
        s = str(n)
        add(s)
        for w in (2, 3, 4, 5, 6):
            add(s.zfill(w))
    except (ValueError, OverflowError):
        pass
    return out


def _find_product(session: Session, code: str) -> Product | None:
    for key in _code_lookup_variants(code):
        p = session.exec(select(Product).where(Product.cod_produto == key)).first()
        if p:
            return p
    return None


def _record_history(session: Session, product_id: int, field: str, old_val, new_val, actor: str) -> None:
    session.add(
        ProductHistory(
            product_id=product_id,
            field_name=field,
            old_value=str(old_val) if old_val is not None else None,
            new_value=str(new_val) if new_val is not None else None,
            changed_by=actor,
        )
    )


def _is_close(a, b) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return math.isclose(float(a), float(b), rel_tol=0, abs_tol=1e-9)


def main() -> int:
    parser = argparse.ArgumentParser(description="Aplica fatores UN/CX e/ou CX/PL de planilha ao cadastro.")
    parser.add_argument("xlsx", type=Path, help="Caminho do .xlsx (ex.: Palete.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Só lista alterações, sem gravar")
    parser.add_argument(
        "--actor",
        default="script:apply_palete_factors",
        help="Identificador gravado em product_history / auditoria",
    )
    args = parser.parse_args()
    path: Path = args.xlsx.expanduser()
    if not path.is_file():
        print(f"Arquivo não encontrado: {path}", file=sys.stderr)
        return 1

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first:
        print("Planilha vazia.", file=sys.stderr)
        return 1

    mapped = [HEADER_ALIASES.get(_norm_header(str(h or ""))) for h in first]
    try:
        i_cod = mapped.index("cod_produto")
    except ValueError:
        print("Cabeçalho de código do produto não encontrado.", file=sys.stderr)
        print(f"Mapeado: {list(zip(first, mapped))}", file=sys.stderr)
        return 1

    i_cx: int | None = None
    i_pl: int | None = None
    try:
        i_cx = mapped.index("conversion_factor")
    except ValueError:
        pass
    try:
        i_pl = mapped.index("pallet_conversion_factor")
    except ValueError:
        pass

    if i_cx is None and i_pl is None:
        print(
            "Nenhuma coluna de fator encontrada (UN por 1 CX ou CX por 1 PL). "
            "Inclua cabeçalhos reconhecidos pelo cadastro, ex.: "
            "'Fator de conversão (UN por 1 CX)' e/ou 'Fator de conversão (CX por 1 PL)'.",
            file=sys.stderr,
        )
        print(f"Mapeado: {list(zip(first, mapped))}", file=sys.stderr)
        return 1

    # cod -> { 'conversion_factor': float?, 'pallet_conversion_factor': float? }
    merged: dict[str, dict[str, float]] = {}
    order: list[str] = []
    dup: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_cod = row[i_cod] if i_cod < len(row) else None
        cod = _norm_cod_cell(raw_cod)
        if not cod:
            continue

        delta: dict[str, float] = {}
        if i_cx is not None:
            raw_cx = row[i_cx] if i_cx < len(row) else None
            v = _factor_un_per_cx_from_cell(raw_cx)
            if v is not None:
                delta["conversion_factor"] = v
        if i_pl is not None:
            raw_pl = row[i_pl] if i_pl < len(row) else None
            v = _factor_cx_per_pl_from_cell(raw_pl)
            if v is not None:
                delta["pallet_conversion_factor"] = v

        if not delta:
            print(f"Aviso: sem fator válido para código {cod!r} — linha ignorada.", file=sys.stderr)
            continue

        if cod in merged:
            dup.append(cod)
        elif cod not in merged:
            order.append(cod)
        base = merged.setdefault(cod, {})
        base.update(delta)

    if dup:
        print(f"Aviso: códigos repetidos na planilha (último valor vence por campo): {sorted(set(dup))}", file=sys.stderr)

    session: Session = SessionLocal()
    try:
        _ensure_product_pallet_conversion_factor_column(session)

        pending: list[tuple[Product, list[tuple[str, float | None, float]]]] = []
        missing: list[str] = []
        noop = 0

        for cod in order:
            plan = merged.get(cod) or {}
            p = _find_product(session, cod)
            if not p:
                missing.append(cod)
                continue

            changes: list[tuple[str, float | None, float]] = []
            if "conversion_factor" in plan:
                new_v = plan["conversion_factor"]
                old_v = p.conversion_factor
                if not _is_close(old_v, new_v):
                    changes.append(("conversion_factor", old_v, new_v))
            if "pallet_conversion_factor" in plan:
                new_v = plan["pallet_conversion_factor"]
                old_v = p.pallet_conversion_factor
                if not _is_close(old_v, new_v):
                    changes.append(("pallet_conversion_factor", old_v, new_v))

            if not changes:
                noop += 1
            else:
                pending.append((p, changes))

        print(
            f"Planilha: {len(merged)} código(s) com fator | sem mudança: {noop} | "
            f"a atualizar: {len(pending)} | não encontrados: {len(missing)}",
        )
        if missing:
            print(f"Produtos não encontrados no banco ({len(missing)}): {missing[:30]}{'…' if len(missing) > 30 else ''}")

        if args.dry_run:
            for p, changes in pending[:50]:
                parts = [f"{f} {o!r} -> {n!r}" for f, o, n in changes]
                print(f"  {p.cod_produto!r}: {'; '.join(parts)} | {p.cod_grup_descricao[:50]!r}")
            if len(pending) > 50:
                print(f"  … e mais {len(pending) - 50} produto(s).")
            return 0

        updated_products = 0
        for p, changes in pending:
            for field, old_v, new_v in changes:
                _record_history(session, p.id or 0, field, old_v, new_v, args.actor)
                setattr(p, field, new_v)
            p.updated_at = utcnow()
            session.add(p)
            updated_products += 1

        if updated_products:
            _safe_log_change(
                session,
                "products",
                0,
                "bulk_product_factors_xlsx",
                args.actor,
                {
                    "file": str(path),
                    "products_touched": updated_products,
                    "missing_codes": missing,
                    "columns": {"conversion_factor": i_cx is not None, "pallet_conversion_factor": i_pl is not None},
                },
            )
            session.commit()
        print(f"Gravado: {updated_products} produto(s) com alteração(ões).")
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
