import hashlib
import logging
import re
import uuid
from collections import defaultdict
from itertools import groupby
from datetime import date, datetime, time, timezone, timedelta
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import desc, func
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, SQLModel, select

from app.api.deps import require_roles, require_stock_analysis_access
from app.db.session import engine, get_session
from app.models import (
    ChangeLog,
    InventoryImport,
    InventoryImportItem,
    MateCouroTrocaLog,
    Product,
    RecountSignal,
    User,
    ValidityLine,
)
from app.services.inventory_txt_parse import resolve_saldo_fisico_caixa_unidade
from app.services.validity_lines_schema import ensure_validity_lines_structures

router = APIRouter(prefix="/audit", tags=["audit"])
logger = logging.getLogger(__name__)
_BR = ZoneInfo("America/Sao_Paulo")

# Janela ao buscar a última diferença contagem × TXT antes do dia da análise (sem base de troca).
_STOCK_ANALYSIS_PREV_DIFF_LOOKBACK_DAYS = 120


def _ensure_validity_lines_table() -> None:
    """Garante DDL (Railway/deploy sem alembic ou migração pendente). Idempotente."""
    try:
        ensure_validity_lines_structures()
    except SQLAlchemyError:
        logger.exception("Falha ao garantir tabela validity_lines")
        raise


def _validity_observed_at_iso(dt: datetime | None) -> str | None:
    if dt is None:
        return None
    try:
        if dt.tzinfo is None:
            u = dt.replace(tzinfo=timezone.utc)
        else:
            u = dt.astimezone(timezone.utc)
        return u.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    except Exception:
        return None


def _parse_iso_date_arg(value: str | None) -> date | None:
    if not value or not str(value).strip():
        return None
    try:
        return date.fromisoformat(str(value).strip()[:10])
    except ValueError:
        return None


def _brazil_date_from_observed_at(observed_at: str | None) -> date | None:
    if not observed_at:
        return None
    try:
        s = str(observed_at).strip()
        if s.endswith("Z"):
            s = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_BR).date()
    except Exception:
        return None


def _parse_and_validate_observed_at(observed_at: str) -> tuple[datetime, date]:
    """Valida ISO8601; retorna instante UTC-normalizado e data em America/Sao_Paulo."""
    try:
        s = str(observed_at).strip()
        if s.endswith("Z"):
            s = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        br_date = dt.astimezone(_BR).date()
    except Exception as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"observed_at invalido: {exc}",
        ) from exc
    now_br = datetime.now(timezone.utc).astimezone(_BR)
    if br_date > now_br.date():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Data da operacao (America/Sao_Paulo) nao pode ser futura.",
        )
    if dt > datetime.now(timezone.utc) + timedelta(hours=2):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="observed_at muito no futuro em relacao ao servidor.",
        )
    return dt, br_date


class CountEventInput(BaseModel):
    client_event_id: str = Field(min_length=8, max_length=100)
    item_code: str = Field(min_length=1, max_length=120)
    quantity: int = Field(ge=-500000, le=500000)
    observed_at: str
    device_name: str | None = Field(default=None, max_length=120)


class CountEventsPayload(BaseModel):
    events: list[CountEventInput]


class BreakEventInput(BaseModel):
    client_event_id: str = Field(min_length=8, max_length=100)
    item_code: str = Field(min_length=1, max_length=120)
    quantity: int = Field(ge=-500_000, le=500_000)
    observed_at: str
    device_name: str | None = Field(default=None, max_length=120)
    reason: str = Field(min_length=1, max_length=200)
    operational_date: str | None = Field(
        default=None,
        description="YYYY-MM-DD (America/Sao_Paulo). Se omitido, deriva de observed_at.",
    )


class BreakEventsPayload(BaseModel):
    events: list[BreakEventInput]


BREAK_BULK_DELETE_DAY_PHRASE = "APAGAR TODAS AS QUEBRAS DO DIA"


class BreakEventsBulkDeleteBody(BaseModel):
    """Remove lançamentos de quebra no ChangeLog (visível para todos após commit)."""

    operational_date: str = Field(min_length=10, max_length=10)
    cod_produtos: list[str] = Field(default_factory=list)
    confirm_phrase: str = Field(default="", max_length=200)


def _break_payload_operational_date(payload: dict) -> date | None:
    od = payload.get("operational_date")
    if od:
        try:
            return date.fromisoformat(str(od)[:10])
        except ValueError:
            return None
    return _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))


def _collect_break_event_logs_for_day_and_codes(
    session: Session,
    operational_date: date,
    cod_filters: set[str] | None,
) -> list[ChangeLog]:
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    out: list[ChangeLog] = []
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        d_op = _break_payload_operational_date(payload)
        if d_op is None or d_op != operational_date:
            continue
        if cod_filters is not None:
            code = _normalize_numeric_product_code_key(str(payload.get("item_code") or ""))
            if not code or code not in cod_filters:
                continue
        out.append(log)
    return out


class ValidityEventInput(BaseModel):
    client_event_id: str = Field(min_length=8, max_length=100)
    cod_produto: str = Field(min_length=1, max_length=120)
    expiration_date: date
    quantity_un: int = Field(default=0, ge=0, le=500_000)
    quantity_cx: int = Field(default=0, ge=0, le=500_000)
    lot_code: str | None = Field(default=None, max_length=80)
    note: str | None = Field(default=None, max_length=500)
    observed_at: str
    device_name: str | None = Field(default=None, max_length=120)


class ValidityEventsPayload(BaseModel):
    events: list[ValidityEventInput]
    reference_date: str | None = Field(
        default=None,
        description="YYYY-MM-DD da importação TXT usada para teto de UN/CX (mesmo filtro da contagem).",
    )


def _un_balance_map_for_validity(session: Session, reference_date: date | None) -> dict[str, int]:
    """Saldo UN agregado por código (última importação para a data informada)."""
    current_import = None
    if reference_date is not None:
        current_import = session.exec(
            select(InventoryImport)
            .where(InventoryImport.reference_date == reference_date)
            .order_by(InventoryImport.imported_at.desc())
            .limit(1)
        ).first()
    else:
        current_import = session.exec(
            select(InventoryImport).order_by(InventoryImport.imported_at.desc()).limit(1)
        ).first()

    out: dict[str, int] = {}
    if not current_import:
        return out

    import_items = list(
        session.exec(
            select(InventoryImportItem).where(InventoryImportItem.inventory_import_id == current_import.id)
        ).all()
    )
    for item in import_items:
        code = _normalize_item_code(item.cod_produto)
        if not code:
            continue
        _cx, un = _extract_import_quantities(item.metrics if isinstance(item.metrics, dict) else None)
        out[code] = out.get(code, 0) + int(un)
    return out


def _cx_balance_map_for_validity(session: Session, reference_date: date | None) -> dict[str, int]:
    """Saldo CX agregado por código (última importação para a data informada)."""
    current_import = None
    if reference_date is not None:
        current_import = session.exec(
            select(InventoryImport)
            .where(InventoryImport.reference_date == reference_date)
            .order_by(InventoryImport.imported_at.desc())
            .limit(1)
        ).first()
    else:
        current_import = session.exec(
            select(InventoryImport).order_by(InventoryImport.imported_at.desc()).limit(1)
        ).first()

    out: dict[str, int] = {}
    if not current_import:
        return out

    import_items = list(
        session.exec(
            select(InventoryImportItem).where(InventoryImportItem.inventory_import_id == current_import.id)
        ).all()
    )
    for item in import_items:
        code = _normalize_item_code(item.cod_produto)
        if not code:
            continue
        cx, _un = _extract_import_quantities(item.metrics if isinstance(item.metrics, dict) else None)
        out[code] = out.get(code, 0) + int(cx)
    return out


def _normalize_item_code(value: str | None) -> str:
    raw = (value or "").strip().upper()
    raw = re.sub(r"\s+", " ", raw)
    raw = re.sub(r"\s*\[(UN|CX)\]\s*$", "", raw)
    return raw


def _normalize_numeric_product_code_key(value: str | None) -> str:
    """Unifica chaves só numéricas (ex.: 010 → 10) em quebra e pendente Mate couro."""
    c = _normalize_item_code(value)
    if not c:
        return ""
    if c.isdigit():
        return str(int(c))
    return c


def _mate_troca_cod_preferred_over_alias(cod_raw: str, canon: str) -> bool:
    """True se cod_raw já está na forma canônica (ex. 10) e não em alias (ex. 010)."""
    raw = _normalize_item_code(cod_raw)
    if not raw or _normalize_numeric_product_code_key(raw) != canon:
        return False
    if raw.isdigit():
        return raw == str(int(raw))
    return raw == canon


def _format_first_second_name(full_name: str | None, fallback_login: str | None) -> str:
    """Primeiro nome completo; se houver segundo token, só a inicial e ponto (ex.: Felipe R.). Senão login."""
    fn = (full_name or "").strip()
    if fn:
        parts = fn.split()
        if len(parts) >= 2:
            initial = (parts[1] or "")[:1]
            if initial:
                return f"{parts[0]} {initial.upper()}."
            return parts[0]
        if parts:
            return parts[0]
    fb = (fallback_login or "").strip()
    return fb if fb else "—"


def _display_name_map_for_logins(session: Session, logins: set[str]) -> dict[str, str]:
    """username (login) → rótulo para exibição (primeiro nome + inicial do segundo)."""
    out: dict[str, str] = {}
    clean = {x.strip() for x in logins if x and str(x).strip()}
    if not clean:
        return out
    users = list(session.exec(select(User).where(User.username.in_(list(clean)))).all())
    for u in users:
        lu = (u.username or "").strip()
        if lu:
            out[lu] = _format_first_second_name(u.full_name, u.username)
    for login in clean:
        if login not in out:
            out[login] = login
    return out


def _actor_csv_to_display_labels(actor_csv: str | None, name_map: dict[str, str]) -> str | None:
    if actor_csv is None or not str(actor_csv).strip():
        return None
    parts = [p.strip() for p in re.split(r",\s*", str(actor_csv).strip()) if p.strip()]
    if not parts:
        return None
    return ", ".join(name_map.get(p, p) for p in parts)


def _extract_count_type(value: str | None) -> str:
    raw = (value or "").strip().upper()
    if raw.endswith("[UN]"):
        return "unidade"
    return "caixa"


def _aggregate_count_events_by_code(
    session: Session,
    count_on_date: date | None = None,
) -> dict[str, dict[str, int]]:
    """Soma CX/UN por produto (ChangeLog). Se count_on_date, filtra pela data observada em America/Sao_Paulo."""
    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )
    counted_by_code: dict[str, dict[str, int]] = {}
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        if count_on_date is not None:
            od = payload.get("observed_at")
            br_d = _brazil_date_from_observed_at(str(od) if od is not None else "")
            if br_d is None or br_d != count_on_date:
                continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_item_code(item_code)
        if not code:
            continue
        count_type = _extract_count_type(item_code)
        qty_raw = payload.get("quantity", 0)
        try:
            qty = int(qty_raw)
        except Exception:
            qty = 0
        if code not in counted_by_code:
            counted_by_code[code] = {"caixa": 0, "unidade": 0}
        counted_by_code[code][count_type] = counted_by_code[code].get(count_type, 0) + qty
    return counted_by_code


def _dimensions_with_count_touch_on_date(session: Session, count_on_date: date) -> dict[str, set[str]]:
    """Por código, quais dimensões (caixa / unidade) tiveram ao menos um evento no dia (inclui confirmação qty 0)."""
    by_code: dict[str, set[str]] = defaultdict(set)
    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
        if br_d is None or br_d != count_on_date:
            continue
        raw_item = str(payload.get("item_code") or "")
        code = _normalize_item_code(raw_item)
        if not code:
            continue
        ct = _extract_count_type(raw_item)
        by_code[code].add(ct)
    return dict(by_code)


def _load_imported_by_code_for_reference_date(session: Session, ref_date: date) -> dict[str, dict] | None:
    """Itens agregados da última importação TXT cuja data de referência é `ref_date`."""
    current_import = session.exec(
        select(InventoryImport)
        .where(InventoryImport.reference_date == ref_date)
        .order_by(InventoryImport.imported_at.desc())
        .limit(1)
    ).first()
    if not current_import:
        return None
    imported_by_code: dict[str, dict] = {}
    import_items = list(
        session.exec(
            select(InventoryImportItem).where(InventoryImportItem.inventory_import_id == current_import.id)
        ).all()
    )
    for item in import_items:
        code = _normalize_item_code(item.cod_produto)
        if not code:
            continue
        import_caixa, import_unidade = _extract_import_quantities(
            item.metrics if isinstance(item.metrics, dict) else None
        )
        if code not in imported_by_code:
            imported_by_code[code] = {
                "cod_produto": code,
                "descricao": item.descricao or "",
                "import_caixa": 0,
                "import_unidade": 0,
            }
        imported_by_code[code]["import_caixa"] += import_caixa
        imported_by_code[code]["import_unidade"] += import_unidade
        if not imported_by_code[code]["descricao"] and item.descricao:
            imported_by_code[code]["descricao"] = item.descricao
    return imported_by_code


def _build_count_totals_and_activity_dates_before(
    session: Session,
    count_day: date,
    min_day: date,
) -> tuple[dict[date, dict[str, dict[str, int]]], dict[str, set[date]]]:
    """Totais CX/UN por (dia BR, código) e dias com lançamento, apenas para min_day <= dia < count_day."""
    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )
    totals: dict[date, dict[str, dict[str, int]]] = {}
    code_dates: dict[str, set[date]] = defaultdict(set)
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        raw_item = str(payload.get("item_code") or "")
        br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
        if br_d is None or br_d >= count_day or br_d < min_day:
            continue
        code = _normalize_item_code(raw_item)
        if not code:
            continue
        count_type = _extract_count_type(raw_item)
        qty_raw = payload.get("quantity", 0)
        try:
            qty = int(qty_raw)
        except Exception:
            qty = 0
        if br_d not in totals:
            totals[br_d] = {}
        if code not in totals[br_d]:
            totals[br_d][code] = {"caixa": 0, "unidade": 0}
        totals[br_d][code][count_type] = totals[br_d][code].get(count_type, 0) + qty
        code_dates[code].add(br_d)
    return totals, dict(code_dates)


def _resolve_previous_operational_count_diff(
    session: Session,
    item_code: str,
    count_day: date,
    min_day: date,
    totals_by_date: dict[date, dict[str, dict[str, int]]],
    code_activity_dates: dict[str, set[date]],
    import_cache: dict[date, dict[str, dict] | None],
) -> tuple[date | None, int, int]:
    """
    Último dia < count_day em que o código teve lançamento de contagem e existia TXT com esse código:
    diferença = contagem do dia − saldo TXT daquele dia (sem troca).
    """
    code = _normalize_item_code(item_code)
    if not code:
        return None, 0, 0
    days = code_activity_dates.get(code)
    if not days:
        return None, 0, 0
    for d in sorted(days, reverse=True):
        if d >= count_day:
            continue
        if d < min_day:
            break
        if d not in import_cache:
            import_cache[d] = _load_imported_by_code_for_reference_date(session, d)
        imp_map = import_cache[d]
        if imp_map is None:
            continue
        imp_rec = imp_map.get(code)
        if not imp_rec:
            continue
        day_bucket = totals_by_date.get(d) or {}
        counted = day_bucket.get(code) or {"caixa": 0, "unidade": 0}
        icx = int(imp_rec.get("import_caixa", 0))
        iun = int(imp_rec.get("import_unidade", 0))
        ccx = int(counted.get("caixa", 0))
        cun = int(counted.get("unidade", 0))
        return d, ccx - icx, cun - iun
    return None, 0, 0


def _enrich_stock_analysis_rows_previous_operational_diff(
    session: Session,
    rows: list[dict],
    count_day: date,
) -> None:
    if not rows:
        return
    min_day = count_day - timedelta(days=_STOCK_ANALYSIS_PREV_DIFF_LOOKBACK_DAYS)
    totals_by_date, code_dates = _build_count_totals_and_activity_dates_before(session, count_day, min_day)
    import_cache: dict[date, dict[str, dict] | None] = {}
    for row in rows:
        d_prev, dcx, dun = _resolve_previous_operational_count_diff(
            session,
            str(row.get("cod_produto") or ""),
            count_day,
            min_day,
            totals_by_date,
            code_dates,
            import_cache,
        )
        if d_prev is not None:
            row["previous_difference_date"] = d_prev.isoformat()
            row["previous_difference_caixa"] = dcx
            row["previous_difference_unidade"] = dun
        else:
            row["previous_difference_date"] = None
            row["previous_difference_caixa"] = None
            row["previous_difference_unidade"] = None


def _parse_payload_observed_at_utc(od: object) -> datetime | None:
    if od is None:
        return None
    try:
        s = str(od).strip()
        if not s:
            return None
        if s.endswith("Z"):
            s = s.replace("Z", "+00:00")
        dt = datetime.fromisoformat(s)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt
    except Exception:
        return None


def _count_day_activity_meta(session: Session, count_on_date: date) -> dict:
    """Atores e intervalo de observed_at (UTC ISO) dos eventos do dia operacional em SP."""
    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )
    actors: set[str] = set()
    observed_utc: list[datetime] = []
    event_count = 0
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("observed_at")
        br_d = _brazil_date_from_observed_at(str(od) if od is not None else "")
        if br_d is None or br_d != count_on_date:
            continue
        event_count += 1
        a = (log.actor or "").strip()
        if a:
            actors.add(a)
        pdt = _parse_payload_observed_at_utc(od)
        if pdt:
            observed_utc.append(pdt)
    first_iso: str | None = None
    last_iso: str | None = None
    if observed_utc:
        mn = min(observed_utc)
        mx = max(observed_utc)
        first_iso = mn.replace(microsecond=0).isoformat().replace("+00:00", "Z")
        last_iso = mx.replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return {
        "count_date": count_on_date.isoformat(),
        "actors": sorted(actors),
        "first_observed_at": first_iso,
        "last_observed_at": last_iso,
        "event_count": event_count,
    }


def _last_count_snapshot_per_code(session: Session) -> dict[str, dict]:
    """
    Por produto: totais de CX/UN do dia da última contagem registrada no sistema
    (maior data America/Sao_Paulo em observed_at entre eventos de contagem).
    """
    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )
    per_day: dict[tuple[str, date], dict[str, int]] = {}
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("observed_at")
        br_d = _brazil_date_from_observed_at(str(od) if od is not None else "")
        if br_d is None:
            continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_item_code(item_code)
        if not code:
            continue
        count_type = _extract_count_type(item_code)
        qty_raw = payload.get("quantity", 0)
        try:
            qty = int(qty_raw)
        except Exception:
            qty = 0
        key = (code, br_d)
        if key not in per_day:
            per_day[key] = {"caixa": 0, "unidade": 0}
        per_day[key][count_type] = per_day[key].get(count_type, 0) + qty

    by_code: dict[str, tuple[date, dict[str, int]]] = {}
    for (code, br_d), rec in per_day.items():
        if code not in by_code or br_d > by_code[code][0]:
            by_code[code] = (br_d, dict(rec))

    out: dict[str, dict] = {}
    for code, (best_d, rec) in by_code.items():
        out[code] = {
            "caixa": int(rec.get("caixa", 0)),
            "unidade": int(rec.get("unidade", 0)),
            "count_date": best_d.isoformat(),
        }
    return out


def _extract_import_quantities(metrics: dict | list | None) -> tuple[int, int]:
    """CX/UN a partir do JSON do item (preferência: caixa/unidade gravados na importação)."""
    if metrics is None:
        return 0, 0
    if isinstance(metrics, dict):
        caixa = metrics.get("caixa")
        unidade = metrics.get("unidade")
        if caixa is not None and unidade is not None:
            try:
                return int(caixa), int(unidade)
            except (TypeError, ValueError):
                pass
        nt = metrics.get("numeric_tail")
        raw = metrics.get("raw")
        nt_str = nt if isinstance(nt, str) else None
        raw_list = [str(x) for x in raw] if isinstance(raw, list) else []
        return resolve_saldo_fisico_caixa_unidade(nt_str, raw_list)
    if isinstance(metrics, list):
        return resolve_saldo_fisico_caixa_unidade(None, [str(x) for x in metrics])
    return 0, 0


@router.get("/import-balances")
def import_balances(
    reference_date: str | None = Query(default=None),
    only_active: bool = Query(default=True, alias="only_active"),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """
    Saldo CX/UN por produto conforme importação TXT (mesma base da análise de contagem).
    has_txt_import=false quando não há arquivo (fallback catálogo 0/0) — a UI não deve validar “bateu”.
    """
    from app.api.routes.products import _catalog_status_is_ativo_clause

    current_import = None
    if reference_date:
        current_import = session.exec(
            select(InventoryImport)
            .where(InventoryImport.reference_date == reference_date)
            .order_by(InventoryImport.imported_at.desc())
            .limit(1)
        ).first()
    else:
        current_import = session.exec(
            select(InventoryImport).order_by(InventoryImport.imported_at.desc()).limit(1)
        ).first()

    use_catalog_fallback = False
    imported_by_code: dict[str, dict] = {}
    import_meta: dict | None = None

    if current_import:
        import_items = list(
            session.exec(
                select(InventoryImportItem).where(InventoryImportItem.inventory_import_id == current_import.id)
            ).all()
        )
        for item in import_items:
            code = _normalize_item_code(item.cod_produto)
            if not code:
                continue
            import_caixa, import_unidade = _extract_import_quantities(
                item.metrics if isinstance(item.metrics, dict) else None
            )
            if code not in imported_by_code:
                imported_by_code[code] = {
                    "cod_produto": code,
                    "descricao": item.descricao or "",
                    "import_caixa": 0,
                    "import_unidade": 0,
                }
            imported_by_code[code]["import_caixa"] += import_caixa
            imported_by_code[code]["import_unidade"] += import_unidade
            if not imported_by_code[code]["descricao"] and item.descricao:
                imported_by_code[code]["descricao"] = item.descricao
        import_meta = {
            "id": current_import.id,
            "reference_date": current_import.reference_date,
            "file_name": current_import.file_name,
            "imported_at": current_import.imported_at,
            "total_products": current_import.total_products,
            "created_products": current_import.created_products,
        }
    elif only_active:
        use_catalog_fallback = True
        prods = list(
            session.exec(
                select(Product.cod_produto, Product.cod_grup_descricao)
                .where(_catalog_status_is_ativo_clause())
                .order_by(Product.cod_grup_descricao)
            ).all()
        )
        for cod_raw, desc in prods:
            code = _normalize_item_code(cod_raw)
            if not code:
                continue
            imported_by_code[code] = {
                "cod_produto": code,
                "descricao": (desc or "").strip(),
                "import_caixa": 0,
                "import_unidade": 0,
            }
        ref_label = (reference_date or "").strip()
        file_note = (
            f"Sem importação TXT para {ref_label} — saldo 0 CX / 0 UN (somente ativos)"
            if ref_label
            else "Sem importação TXT — saldo 0 CX / 0 UN (somente produtos ativos)"
        )
        import_meta = {
            "id": None,
            "reference_date": reference_date or None,
            "file_name": file_note,
            "imported_at": None,
            "total_products": len(imported_by_code),
            "created_products": 0,
        }
    else:
        return {"has_txt_import": False, "import": None, "balances": {}}

    has_txt_import = bool(current_import) and not use_catalog_fallback

    balances: dict[str, dict[str, int]] = {}
    for code, rec in imported_by_code.items():
        balances[code] = {
            "import_caixa": int(rec.get("import_caixa", 0)),
            "import_unidade": int(rec.get("import_unidade", 0)),
        }

    return {
        "has_txt_import": has_txt_import,
        "import": import_meta,
        "balances": balances,
    }


@router.get("/last-count-per-product")
def last_count_per_product(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """
    Última contagem consolidada por produto (ChangeLog): CX/UN do dia mais recente
    em que houve lançamento de contagem para aquele código.
    """
    balances = _last_count_snapshot_per_code(session)
    return {"balances": balances}


@router.get("/count-server-totals")
def count_server_totals(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
    count_date: str | None = Query(
        default=None,
        description="YYYY-MM-DD: soma apenas eventos cuja data observada (America/Sao_Paulo) é esse dia. Padrão: hoje em SP.",
    ),
) -> dict:
    """
    Totais CX/UN já sincronizados no servidor (ChangeLog), somando conferentes.
    Filtrado por dia operacional (America/Sao_Paulo) para não misturar contagens de dias anteriores.
    """
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    d = _parse_iso_date_arg(count_date) if count_date else br_today
    if d is None:
        d = br_today
    counted_by_code = _aggregate_count_events_by_code(session, count_on_date=d)
    balances: dict[str, dict[str, int]] = {}
    for code, rec in counted_by_code.items():
        balances[code] = {
            "caixa": int(rec.get("caixa", 0)),
            "unidade": int(rec.get("unidade", 0)),
        }
    meta = _count_day_activity_meta(session, d)
    return {"balances": balances, "meta": meta}


def _compute_stock_analysis(
    session: Session,
    import_id: int | None,
    reference_date: str | None,
    only_diff: bool,
    only_active_products: bool,
    limit: int,
) -> dict:
    from app.api.routes.products import _catalog_status_is_ativo_clause

    current_import = None
    if import_id is not None:
        current_import = session.get(InventoryImport, import_id)
    elif reference_date:
        # Busca importação pela data de referência exata
        current_import = session.exec(
            select(InventoryImport)
            .where(InventoryImport.reference_date == reference_date)
            .order_by(InventoryImport.imported_at.desc())
            .limit(1)
        ).first()
    else:
        current_import = session.exec(
            select(InventoryImport).order_by(InventoryImport.imported_at.desc()).limit(1)
        ).first()

    use_catalog_fallback = False
    imported_by_code: dict[str, dict] = {}
    import_meta: dict | None = None

    if current_import:
        import_items = list(
            session.exec(
                select(InventoryImportItem).where(InventoryImportItem.inventory_import_id == current_import.id)
            ).all()
        )
        for item in import_items:
            code = _normalize_item_code(item.cod_produto)
            if not code:
                continue
            import_caixa, import_unidade = _extract_import_quantities(
                item.metrics if isinstance(item.metrics, dict) else None
            )
            if code not in imported_by_code:
                imported_by_code[code] = {
                    "cod_produto": code,
                    "descricao": item.descricao or "",
                    "import_caixa": 0,
                    "import_unidade": 0,
                }
            imported_by_code[code]["import_caixa"] += import_caixa
            imported_by_code[code]["import_unidade"] += import_unidade
            if not imported_by_code[code]["descricao"] and item.descricao:
                imported_by_code[code]["descricao"] = item.descricao
        import_meta = {
            "id": current_import.id,
            "reference_date": current_import.reference_date,
            "file_name": current_import.file_name,
            "imported_at": current_import.imported_at,
            "total_products": current_import.total_products,
            "created_products": current_import.created_products,
        }
    elif only_active_products:
        # Sem TXT: baseia a análise no catálogo de produtos ativos com saldo 0 CX / 0 UN
        use_catalog_fallback = True
        prods = list(
            session.exec(
                select(Product.cod_produto, Product.cod_grup_descricao)
                .where(_catalog_status_is_ativo_clause())
                .order_by(Product.cod_grup_descricao)
            ).all()
        )
        for cod_raw, desc in prods:
            code = _normalize_item_code(cod_raw)
            if not code:
                continue
            imported_by_code[code] = {
                "cod_produto": code,
                "descricao": (desc or "").strip(),
                "import_caixa": 0,
                "import_unidade": 0,
            }
        ref_label = (reference_date or "").strip()
        file_note = (
            f"Sem importação TXT para {ref_label} — saldo 0 CX / 0 UN (somente ativos)"
            if ref_label
            else "Sem importação TXT — saldo 0 CX / 0 UN (somente produtos ativos)"
        )
        import_meta = {
            "id": None,
            "reference_date": reference_date or None,
            "file_name": file_note,
            "imported_at": None,
            "total_products": len(imported_by_code),
            "created_products": 0,
        }
    else:
        return {
            "import": None,
            "summary": {
                "total_import_items": 0,
                "counted_items": 0,
                "equal_items": 0,
                "divergent_items": 0,
                "missing_in_count": 0,
                "extra_in_count": 0,
            },
            "rows": [],
        }

    active_set: set[str] | None = None
    if only_active_products:
        if use_catalog_fallback:
            active_set = set(imported_by_code.keys())
        else:
            active_rows = session.exec(select(Product.cod_produto).where(_catalog_status_is_ativo_clause())).all()
            active_set = {_normalize_item_code(c) for c in active_rows if c}

    # Contagem comparada ao saldo: apenas eventos do mesmo dia de referência (America/Sao_Paulo).
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    if current_import:
        rd = current_import.reference_date
        count_day = rd if isinstance(rd, date) else _parse_iso_date_arg(str(rd))
        if count_day is None:
            count_day = br_today
    else:
        count_day = _parse_iso_date_arg(reference_date) or br_today
    counted_by_code = _aggregate_count_events_by_code(session, count_on_date=count_day)
    dim_touches_by_code = _dimensions_with_count_touch_on_date(session, count_day)

    all_codes = set(imported_by_code.keys()) | set(counted_by_code.keys())
    rows: list[dict] = []
    equal_items = 0
    divergent_items = 0
    missing_in_count = 0
    extra_in_count = 0

    grupos_by_code: dict[str, str] = {}
    if only_active_products:
        prod_grup = session.exec(
            select(Product.cod_produto, Product.cod_grup_familia, Product.cod_grup_segmento).where(
                _catalog_status_is_ativo_clause()
            )
        ).all()
        for cod_raw, fam, seg in prod_grup:
            c = _normalize_item_code(cod_raw)
            if not c:
                continue
            label = (str(fam or "").strip() or str(seg or "").strip())
            grupos_by_code[c] = label if label else "Sem grupo"

    for code in all_codes:
        if active_set is not None and code not in active_set:
            continue
        imported = imported_by_code.get(code)
        import_caixa = int(imported.get("import_caixa", 0)) if imported else 0
        import_unidade = int(imported.get("import_unidade", 0)) if imported else 0
        counted = counted_by_code.get(code, {"caixa": 0, "unidade": 0})
        counted_caixa = int(counted.get("caixa", 0))
        counted_unidade = int(counted.get("unidade", 0))
        diff_caixa = counted_caixa - import_caixa
        diff_unidade = counted_unidade - import_unidade
        total_diff_abs = abs(diff_caixa) + abs(diff_unidade)
        touches = dim_touches_by_code.get(code, set())
        status = "ok"
        if imported and counted_caixa == 0 and counted_unidade == 0:
            if import_caixa != 0 or import_unidade != 0:
                status = "missing_in_count"
                missing_in_count += 1
            else:
                # TXT 0/0: exige confirmação explícita em CX e UN (evento no dia, pode ser quantidade 0).
                if "caixa" not in touches or "unidade" not in touches:
                    status = "missing_in_count"
                    missing_in_count += 1
        elif not imported and (counted_caixa != 0 or counted_unidade != 0):
            status = "extra_in_count"
            extra_in_count += 1
        elif diff_caixa != 0 or diff_unidade != 0:
            status = "divergent"
            divergent_items += 1
        else:
            equal_items += 1

        if only_diff and status == "ok":
            continue

        rows.append(
            {
                "cod_produto": code,
                "descricao": (imported or {}).get("descricao") or "",
                "import_caixa": import_caixa,
                "import_unidade": import_unidade,
                "counted_caixa": counted_caixa,
                "counted_unidade": counted_unidade,
                "difference_caixa": diff_caixa,
                "difference_unidade": diff_unidade,
                "difference_abs": total_diff_abs,
                "status": status,
                "grupo": grupos_by_code.get(code, "Sem grupo"),
            }
        )

    def _status_rank(st: str) -> int:
        return {"missing_in_count": 0, "divergent": 1, "extra_in_count": 2, "ok": 3}.get(st, 9)

    rows.sort(
        key=lambda r: (
            _status_rank(str(r["status"])),
            -int(r["difference_abs"]),
            (r.get("grupo") or "Sem grupo").lower(),
            str(r["cod_produto"]),
        )
    )
    rows = rows[:limit]
    _enrich_stock_analysis_rows_previous_operational_diff(session, rows, count_day)

    def _len_codes(mapping: dict[str, dict]) -> int:
        if active_set is None:
            return len(mapping)
        return sum(1 for c in mapping if c in active_set)

    return {
        "import": import_meta,
        "summary": {
            "total_import_items": _len_codes(imported_by_code),
            "counted_items": _len_codes(counted_by_code),
            "equal_items": equal_items,
            "divergent_items": divergent_items,
            "missing_in_count": missing_in_count,
            "extra_in_count": extra_in_count,
        },
        "rows": rows,
    }


def _safe_int(value: object) -> int:
    try:
        return int(value)
    except Exception:
        return 0


def _timeline_count_events_for_code(
    session: Session,
    item_code: str,
    count_on_date: date | None = None,
) -> list[dict]:
    code = _normalize_item_code(item_code)
    if not code:
        return []

    count_logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_count",
                ChangeLog.action == "count_event",
            )
        ).all()
    )

    events: list[dict] = []
    for log in count_logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        raw_item_code = str(payload.get("item_code") or "")
        if _normalize_item_code(raw_item_code) != code:
            continue

        observed_at = str(payload.get("observed_at") or "").strip() or None
        if count_on_date is not None:
            br_d = _brazil_date_from_observed_at(observed_at)
            if br_d is None or br_d != count_on_date:
                continue

        changed_at = _validity_observed_at_iso(log.changed_at)
        events.append(
            {
                "log_id": log.id,
                "actor": (log.actor or "").strip() or None,
                "observed_at": observed_at,
                "changed_at": changed_at,
                "device_name": (payload.get("device_name") or "") or None,
                "count_type": _extract_count_type(raw_item_code),
                "quantity_delta": _safe_int(payload.get("quantity", 0)),
                "_sort_observed": observed_at or "",
                "_sort_changed": changed_at or "",
            }
        )

    events.sort(key=lambda e: (e["_sort_observed"], e["_sort_changed"], int(e.get("log_id") or 0)))

    total_caixa = 0
    total_unidade = 0
    history: list[dict] = []
    for event in events:
        prev_caixa = total_caixa
        prev_unidade = total_unidade
        qty = int(event["quantity_delta"])
        if event["count_type"] == "unidade":
            total_unidade += qty
        else:
            total_caixa += qty

        history.append(
            {
                "log_id": event["log_id"],
                "actor": event["actor"],
                "observed_at": event["observed_at"],
                "changed_at": event["changed_at"],
                "device_name": event["device_name"],
                "count_type": event["count_type"],
                "quantity_delta": qty,
                "previous_value": prev_unidade if event["count_type"] == "unidade" else prev_caixa,
                "current_value": total_unidade if event["count_type"] == "unidade" else total_caixa,
                "previous_caixa": prev_caixa,
                "current_caixa": total_caixa,
                "previous_unidade": prev_unidade,
                "current_unidade": total_unidade,
            }
        )

    return history


def _audit_status_label_pt(status: str) -> str:
    return {
        "ok": "OK",
        "missing_in_count": "Sem contagem",
        "extra_in_count": "Só na contagem",
        "divergent": "Divergência",
    }.get(status, status)


def _build_stock_analysis_excel_workbook(
    data: dict,
    emitted_at_br: datetime,
    emitted_by: str,
) -> BytesIO:
    """Monta planilha formatada com cabeçalho, resumo e linhas da análise."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Análise"

    title_font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="1B4332",
        end_color="1B4332",
        fill_type="solid",
    )
    sub_fill = PatternFill(
        start_color="D8F3DC",
        end_color="D8F3DC",
        fill_type="solid",
    )
    thin = Side(style="thin", color="B7B7B7")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center = Alignment(horizontal="center", vertical="center")

    ncols = 14
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    c1 = ws["A1"]
    c1.value = "Análise de Contagem · Estoque"
    c1.font = title_font
    c1.fill = header_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"].value = (
        f"Emitido em: {emitted_at_br.strftime('%d/%m/%Y %H:%M:%S')}  ·  "
        f"Emitido por: {emitted_by}"
    )
    ws["A2"].font = Font(name="Calibri", size=11, bold=True, color="1B4332")
    ws["A2"].fill = sub_fill
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    imp = data.get("import") or {}
    ref_line = ""
    if imp:
        ref = imp.get("reference_date") or "—"
        fname = imp.get("file_name") or "—"
        ref_line = f"Base de saldo: data ref. {ref}  |  {fname}"
    else:
        ref_line = "Base de saldo: não disponível"

    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"].value = ref_line
    ws["A3"].font = Font(name="Calibri", size=10)
    ws["A3"].alignment = wrap

    s = data.get("summary") or {}
    summary_text = (
        f"Resumo — Itens com saldo: {s.get('total_import_items', 0)}  |  "
        f"Com contagem: {s.get('counted_items', 0)}  |  "
        f"Conferidos: {s.get('equal_items', 0)}  |  "
        f"Divergências: {s.get('divergent_items', 0)}  |  "
        f"Sem contagem: {s.get('missing_in_count', 0)}  |  "
        f"Só na contagem: {s.get('extra_in_count', 0)}"
    )
    ws.merge_cells(f"A4:{last_col}4")
    ws["A4"].value = summary_text
    ws["A4"].font = Font(name="Calibri", size=10, italic=True)
    ws["A4"].alignment = wrap

    headers = [
        "Código",
        "Grupo",
        "Produto",
        "Situação",
        "Saldo CX",
        "Saldo UN",
        "Contagem CX",
        "Contagem UN",
        "Dif. CX",
        "Dif. UN",
        "|Dif| total",
        "Últ. dif. (data)",
        "Últ. dif. CX",
        "Últ. dif. UN",
    ]
    start_row = 6
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
        cell.alignment = center

    rows = data.get("rows") or []
    for i, row in enumerate(rows, start=start_row + 1):
        status = _audit_status_label_pt(str(row.get("status") or ""))
        pdd = row.get("previous_difference_date")
        pdd_str = str(pdd)[:10] if pdd else ""
        pdc = row.get("previous_difference_caixa")
        pdu = row.get("previous_difference_unidade")
        vals = [
            row.get("cod_produto") or "",
            row.get("grupo") or "",
            row.get("descricao") or "",
            status,
            int(row.get("import_caixa") or 0),
            int(row.get("import_unidade") or 0),
            int(row.get("counted_caixa") or 0),
            int(row.get("counted_unidade") or 0),
            int(row.get("difference_caixa") or 0),
            int(row.get("difference_unidade") or 0),
            int(row.get("difference_abs") or 0),
            pdd_str,
            int(pdc) if pdc is not None else "",
            int(pdu) if pdu is not None else "",
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.border = border_all
            if j > 4:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    widths = [14, 18, 42, 16, 10, 10, 12, 12, 10, 10, 12, 14, 10, 10]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    ws.freeze_panes = f"A{start_row + 1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ValidityAnalysisExportPayload(BaseModel):
    """JSON enviado pelo front com os mesmos dados exibidos na análise (incl. merge local)."""

    operational_date: str = Field(..., min_length=8, max_length=32)
    brazil_today: str | None = None
    last_sync_display: str | None = None
    executive_lines: list[str] = Field(default_factory=list)
    summary_kpis: dict[str, int] = Field(default_factory=dict)
    rows: list[dict[str, Any]] = Field(default_factory=list)
    history_flat: list[dict[str, Any]] = Field(default_factory=list)


def _validity_xlsx_fill_for_visual(key: str | None) -> PatternFill:
    k = (key or "").strip().lower()
    colors = {
        "expired": "FEE2E2",
        "d30": "FECACA",
        "d60": "FFEDD5",
        "d90": "FEF3C7",
        "d120": "FEF9C3",
        "d150": "DBEAFE",
        "d180": "EDE9FE",
        "ok": "DCFCE7",
        "no_validity": "E2E8F0",
        "no_count": "CCFBF1",
        "oldbase": "F3E8FF",
        "none": "F1F5F9",
    }
    hx = colors.get(k, "F8FAFC")
    return PatternFill(start_color=hx, end_color=hx, fill_type="solid")


def _build_validity_analysis_excel_workbook(
    payload: dict[str, Any],
    emitted_at_br: datetime,
    emitted_by: str,
) -> BytesIO:
    wb = Workbook()
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    thin = Side(style="thin", color="CBD5E1")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center = Alignment(horizontal="center", vertical="center")

    # --- Aba 1: Resumo ---
    ws0 = wb.active
    ws0.title = "Resumo Executivo"
    op_d = str(payload.get("operational_date") or "")[:10]
    today = str(payload.get("brazil_today") or op_d)[:10]
    last_sync = str(payload.get("last_sync_display") or "—")
    sk = payload.get("summary_kpis") or {}
    ws0.merge_cells("A1:F1")
    c1 = ws0["A1"]
    c1.value = "Análise de Validades · Resumo executivo"
    c1.font = title_font
    c1.fill = header_fill
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 26

    ws0.merge_cells("A2:F2")
    ws0["A2"].value = (
        f"Emitido em: {emitted_at_br.strftime('%d/%m/%Y %H:%M:%S')}  ·  "
        f"Emitido por: {emitted_by}  ·  Data operacional: {op_d}  ·  Hoje (BR): {today}"
    )
    ws0["A2"].font = Font(name="Calibri", size=10, bold=True)
    ws0["A2"].alignment = wrap

    r = 4
    pairs = [
        ("Última sincronização (cliente)", last_sync),
        ("Total no filtro", sk.get("filtered_total", "—")),
        ("Com validade", sk.get("with", "—")),
        ("Sem validade", sk.get("without", "—")),
        ("Vencidos", sk.get("expired", "—")),
        ("Até 30d", sk.get("d30", "—")),
        ("Até 60d", sk.get("d60", "—")),
        ("Até 90d", sk.get("d90", "—")),
        ("Até 120d", sk.get("d120", "—")),
        ("Até 150d", sk.get("d150", "—")),
        ("Até 180d", sk.get("d180", "—")),
        ("Sem contagem", sk.get("nocount", "—")),
        ("Base antiga", sk.get("oldbase", "—")),
    ]
    ws0.cell(row=r, column=1, value="Métrica").font = header_font
    ws0.cell(row=r, column=1).fill = header_fill
    ws0.cell(row=r, column=2, value="Valor").font = header_font
    ws0.cell(row=r, column=2).fill = header_fill
    r += 1
    for label, val in pairs:
        ws0.cell(row=r, column=1, value=label).border = border_all
        ws0.cell(row=r, column=2, value=val).border = border_all
        r += 1

    r += 1
    ws0.cell(row=r, column=1, value="Narrativa executiva").font = Font(bold=True)
    r += 1
    for line in payload.get("executive_lines") or []:
        ws0.merge_cells(f"A{r}:F{r}")
        ws0.cell(row=r, column=1, value=str(line)).alignment = wrap
        r += 1

    r += 1
    ws0.merge_cells(f"A{r}:F{r}")
    ws0.cell(row=r, column=1, value="Legenda de cores (faixa)").font = Font(bold=True)
    r += 1
    legend = [
        ("Vencido", "FEE2E2"),
        ("≤30d", "FECACA"),
        ("≤60d", "FFEDD5"),
        ("≤90d", "FEF3C7"),
        ("≤120d", "FEF9C3"),
        ("≤150d", "DBEAFE"),
        ("≤180d", "EDE9FE"),
        (">180d", "DCFCE7"),
        ("Sem validade", "E2E8F0"),
        ("Sem contagem", "CCFBF1"),
        ("Base antiga", "F3E8FF"),
    ]
    for lab, hx in legend:
        ws0.cell(row=r, column=1, value=lab)
        ws0.cell(row=r, column=2, value="").fill = PatternFill(
            start_color=hx, end_color=hx, fill_type="solid"
        )
        r += 1

    for idx in range(1, 7):
        ws0.column_dimensions[get_column_letter(idx)].width = 18 if idx > 2 else 28

    # --- Aba 2: Lista ---
    ws1 = wb.create_sheet("Lista Analítica")
    headers = [
        "Código",
        "Produto",
        "Grupo",
        "Situação",
        "Visual",
        "Próx. venc. (BR)",
        "Dias",
        "Faixa",
        "Qtd c/ val.",
        "Qtd s/ val.",
        "Contagem ref.",
        "Últ. lanç. op.",
        "Resp.",
    ]
    hr = 1
    for j, h in enumerate(headers, start=1):
        cell = ws1.cell(row=hr, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
        cell.alignment = center
    rows = payload.get("rows") or []
    for i, row in enumerate(rows, start=hr + 1):
        vk = str(row.get("visual_key") or "")
        vals = [
            row.get("cod_produto") or "",
            row.get("produto") or "",
            row.get("grupo") or "",
            row.get("situacao") or "",
            vk,
            row.get("proximo_vencimento_br") or "",
            row.get("dias_para_vencer") or "",
            row.get("faixa") or "",
            row.get("qtd_com_validade") or "",
            row.get("qtd_sem_validade") or "",
            row.get("contagem_referencia") or "",
            row.get("ultimo_lancamento_op") or "",
            row.get("responsavel_ultimo") or "",
        ]
        fill = _validity_xlsx_fill_for_visual(vk)
        for j, v in enumerate(vals, start=1):
            cell = ws1.cell(row=i, column=j, value=v)
            cell.border = border_all
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if j <= 5 else "right" if j in (7, 8) else "left", vertical="center")
    last_r = hr + len(rows)
    if rows:
        ws1.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{last_r}"
    ws1.freeze_panes = f"A{hr + 1}"
    widths = [12, 36, 22, 14, 14, 14, 8, 18, 14, 14, 18, 14, 18]
    for idx, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(idx)].width = w

    # --- Aba 3: Críticos ---
    ws2 = wb.create_sheet("Críticos")
    crit_keys = {"expired", "d30", "no_validity", "no_count", "oldbase"}
    crit_rows = [x for x in rows if str(x.get("visual_key") or "") in crit_keys]
    for j, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
    for i, row in enumerate(crit_rows, start=2):
        vk = str(row.get("visual_key") or "")
        vals = [
            row.get("cod_produto") or "",
            row.get("produto") or "",
            row.get("grupo") or "",
            row.get("situacao") or "",
            vk,
            row.get("proximo_vencimento_br") or "",
            row.get("dias_para_vencer") or "",
            row.get("faixa") or "",
            row.get("qtd_com_validade") or "",
            row.get("qtd_sem_validade") or "",
            row.get("contagem_referencia") or "",
            row.get("ultimo_lancamento_op") or "",
            row.get("responsavel_ultimo") or "",
        ]
        fill = _validity_xlsx_fill_for_visual(vk)
        for j, v in enumerate(vals, start=1):
            cell = ws2.cell(row=i, column=j, value=v)
            cell.border = border_all
            cell.fill = fill
    if crit_rows:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(crit_rows) + 1}"
    ws2.freeze_panes = "A2"
    for idx, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(idx)].width = w

    # --- Aba 4: Histórico ---
    ws3 = wb.create_sheet("Histórico")
    hh = ["Código", "Produto", "Vencimento", "CX", "UN", "Observado em", "Responsável", "Dia operacional"]
    for j, h in enumerate(hh, start=1):
        cell = ws3.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border_all
    hist = payload.get("history_flat") or []
    for i, hrow in enumerate(hist, start=2):
        vals = [
            hrow.get("cod_produto") or "",
            hrow.get("produto") or "",
            hrow.get("vencimento") or "",
            hrow.get("cx"),
            hrow.get("un"),
            hrow.get("observado_em") or "",
            hrow.get("responsavel") or "",
            hrow.get("dia_operacional") or "",
        ]
        for j, v in enumerate(vals, start=1):
            cell = ws3.cell(row=i, column=j, value=v)
            cell.border = border_all
    if hist:
        ws3.auto_filter.ref = f"A1:{get_column_letter(len(hh))}{len(hist) + 1}"
    ws3.freeze_panes = "A2"
    w3 = [12, 32, 12, 8, 8, 22, 22, 14]
    for idx, w in enumerate(w3, start=1):
        ws3.column_dimensions[get_column_letter(idx)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.post("/validity-analysis/export.xlsx")
def export_validity_analysis_excel(
    payload: ValidityAnalysisExportPayload,
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> StreamingResponse:
    """
    Gera Excel premium a partir do payload calculado no cliente (alinhado à tela).
    EMAIL_VALIDITY_REPORT_ATTACH: reutilizar o mesmo bytes/body em envio futuro (multipart).
    """
    emitted_at_br = datetime.now(timezone.utc).astimezone(ZoneInfo("America/Sao_Paulo"))
    emitted_by = (user.full_name or "").strip() or (user.username or "—")
    if user.username and user.username not in emitted_by and "@" in (user.username or ""):
        emitted_by = f"{emitted_by} ({user.username})"

    buf = _build_validity_analysis_excel_workbook(
        payload.model_dump(),
        emitted_at_br,
        emitted_by,
    )
    op = str(payload.operational_date or "").strip()[:10] or emitted_at_br.strftime("%Y-%m-%d")
    filename = f"analise_validades_{op}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/stock-analysis")
def stock_analysis(
    import_id: int | None = Query(default=None, ge=1),
    reference_date: str | None = Query(default=None),
    only_diff: bool = Query(default=False),
    only_active_products: bool = Query(default=True, alias="only_active"),
    limit: int = Query(default=500, ge=1, le=5000),
    session: Session = Depends(get_session),
    _: User = Depends(require_stock_analysis_access),
) -> dict:
    return _compute_stock_analysis(
        session,
        import_id=import_id,
        reference_date=reference_date,
        only_diff=only_diff,
        only_active_products=only_active_products,
        limit=limit,
    )


@router.get("/stock-analysis/detail")
def stock_analysis_detail(
    item_code: str = Query(..., min_length=1, max_length=120),
    import_id: int | None = Query(default=None, ge=1),
    reference_date: str | None = Query(default=None),
    only_active_products: bool = Query(default=True, alias="only_active"),
    session: Session = Depends(get_session),
    _: User = Depends(require_stock_analysis_access),
) -> dict:
    normalized = _normalize_item_code(item_code)
    if not normalized:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Codigo do produto invalido para detalhamento.",
        )

    data = _compute_stock_analysis(
        session,
        import_id=import_id,
        reference_date=reference_date,
        only_diff=False,
        only_active_products=only_active_products,
        limit=50_000,
    )

    row = next(
        (
            item
            for item in data.get("rows", [])
            if _normalize_item_code(str(item.get("cod_produto") or "")) == normalized
        ),
        None,
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Produto nao encontrado na analise.")

    import_meta = data.get("import") or {}
    count_day = _parse_iso_date_arg(str(import_meta.get("reference_date") or reference_date or ""))
    if count_day is None:
        count_day = datetime.now(timezone.utc).astimezone(_BR).date()

    history = _timeline_count_events_for_code(session, normalized, count_on_date=count_day)
    actors = sorted({str(h.get("actor") or "").strip() for h in history if str(h.get("actor") or "").strip()})
    devices = sorted(
        {str(h.get("device_name") or "").strip() for h in history if str(h.get("device_name") or "").strip()}
    )

    return {
        "item_code": normalized,
        "count_date": count_day.isoformat(),
        "analysis": row,
        "import": import_meta,
        "history": history,
        "summary": {
            "launches": len(history),
            "actors": actors,
            "devices": devices,
            "last_observed_at": history[-1]["observed_at"] if history else None,
            "last_actor": history[-1]["actor"] if history else None,
        },
    }


@router.get("/stock-analysis/export.xlsx")
def export_stock_analysis_excel(
    import_id: int | None = Query(default=None, ge=1),
    reference_date: str | None = Query(default=None),
    only_diff: bool = Query(default=False),
    only_active_products: bool = Query(default=True, alias="only_active"),
    limit: int = Query(default=20000, ge=1, le=50000),
    session: Session = Depends(get_session),
    user: User = Depends(require_stock_analysis_access),
) -> StreamingResponse:
    """Exporta a mesma análise da tela em Excel com layout e metadados de emissão."""
    data = _compute_stock_analysis(
        session,
        import_id=import_id,
        reference_date=reference_date,
        only_diff=only_diff,
        only_active_products=only_active_products,
        limit=limit,
    )
    emitted_at_br = datetime.now(timezone.utc).astimezone(ZoneInfo("America/Sao_Paulo"))
    emitted_by = (user.full_name or "").strip() or (user.username or "—")
    if user.username and user.username not in emitted_by and "@" in (user.username or ""):
        emitted_by = f"{emitted_by} ({user.username})"

    buf = _build_stock_analysis_excel_workbook(data, emitted_at_br, emitted_by)
    stamp = emitted_at_br.strftime("%Y%m%d_%H%M%S")
    filename = f"analise-contagem_{stamp}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/changes")
def list_changes(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("administrativo", "admin")),
    limit: int = Query(default=100, ge=1, le=500),
) -> list[ChangeLog]:
    statement = select(ChangeLog).order_by(ChangeLog.changed_at.desc()).limit(limit)
    return list(session.exec(statement).all())


@router.post("/count-events")
def ingest_count_events(
    payload: CountEventsPayload,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    synced_ids: list[str] = []
    max_pg_int = 2_147_483_647

    try:
        for event in payload.events:
            # quantity == 0: confirmação explícita (TXT zero na dimensão); persiste para análise / auditoria.
            dt_utc, _br_date = _parse_and_validate_observed_at(event.observed_at)
            obs_stored = dt_utc.replace(microsecond=0).isoformat().replace("+00:00", "Z")

            event_hash = hashlib.sha256(event.client_event_id.encode("utf-8")).hexdigest()
            # Mantem entity_id dentro do limite do INTEGER do PostgreSQL.
            entity_id = (int(event_hash[:16], 16) % max_pg_int) + 1

            existing = session.exec(
                select(ChangeLog).where(
                    ChangeLog.entity_name == "stock_count",
                    ChangeLog.action == "count_event",
                    ChangeLog.entity_id == entity_id,
                )
            ).first()
            if existing:
                synced_ids.append(event.client_event_id)
                continue

            log = ChangeLog(
                entity_name="stock_count",
                entity_id=entity_id,
                action="count_event",
                actor=user.username,
                changed_at=datetime.now(timezone.utc),
                payload={
                    "client_event_id": event.client_event_id,
                    "item_code": event.item_code,
                    "quantity": event.quantity,
                    "observed_at": obs_stored,
                    "device_name": event.device_name,
                },
            )
            session.add(log)
            synced_ids.append(event.client_event_id)

        session.commit()
        return {"received": len(payload.events), "synced": len(synced_ids), "synced_ids": synced_ids}
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Falha ao sincronizar eventos de contagem")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Falha ao sincronizar contagem: {exc}",
        )


def _aggregate_break_events_by_code(session: Session, operational_date: date) -> dict[str, dict[str, int]]:
    """Soma CX/UN de quebra por produto (ChangeLog), filtrando pelo dia operacional."""
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    out: dict[str, dict[str, int]] = {}
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("operational_date")
        if od:
            try:
                d = date.fromisoformat(str(od)[:10])
            except ValueError:
                continue
            if d != operational_date:
                continue
        else:
            br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
            if br_d is None or br_d != operational_date:
                continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_numeric_product_code_key(item_code)
        if not code:
            continue
        try:
            qty = int(payload.get("quantity", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue
        ct = _extract_count_type(item_code)
        if code not in out:
            out[code] = {"caixa": 0, "unidade": 0}
        out[code][ct] = out[code].get(ct, 0) + qty
    return out


def _aggregate_break_events_all_time_by_code(session: Session) -> dict[str, dict[str, int]]:
    """Soma CX/UN de quebra por produto em todo o histórico ChangeLog (break_event)."""
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    out: dict[str, dict[str, int]] = {}
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        item_code = str(payload.get("item_code") or "")
        code = _normalize_numeric_product_code_key(item_code)
        if not code:
            continue
        try:
            qty = int(payload.get("quantity", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue
        ct = _extract_count_type(item_code)
        if code not in out:
            out[code] = {"caixa": 0, "unidade": 0}
        out[code][ct] = out[code].get(ct, 0) + qty
    return out


def _mate_couro_allowed_product_codes(session: Session) -> set[str]:
    """Códigos canônicos (numéricos) dos produtos da CIA Mate couro."""
    rows = list(
        session.exec(select(Product.cod_produto).where(Product.cod_grup_cia == MATE_COURO_CIA)).all()
    )
    allowed: set[str] = set()
    for r in rows:
        c = _normalize_numeric_product_code_key(str(r or ""))
        if c:
            allowed.add(c)
    return allowed


def _normalize_bi_quebras_cia_scope(raw: str | None) -> str:
    scope = (raw or "").strip().lower()
    if scope in ("", BI_QUEBRAS_CIA_SCOPE_ALL, "todos", "all"):
        return BI_QUEBRAS_CIA_SCOPE_ALL
    if scope in (BI_QUEBRAS_CIA_SCOPE_MATE_COURO, "mate_couro", "mate couro", "mate"):
        return BI_QUEBRAS_CIA_SCOPE_MATE_COURO
    if scope in (BI_QUEBRAS_CIA_SCOPE_OUTRAS, "outras-cias", "outras_cias", "other"):
        return BI_QUEBRAS_CIA_SCOPE_OUTRAS
    raise HTTPException(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        detail="Filtro de CIA invalido para BI de quebras.",
    )


def _normalize_bi_quebras_cia_label(raw: str | None) -> str:
    return (raw or "").strip() or "Sem CIA"


def _is_mate_couro_cia(raw: str | None) -> bool:
    return _normalize_bi_quebras_cia_label(raw).casefold() == MATE_COURO_CIA.casefold()


def _matches_bi_quebras_cia_scope(raw_cia: str | None, scope: str) -> bool:
    normalized_scope = _normalize_bi_quebras_cia_scope(scope)
    if normalized_scope == BI_QUEBRAS_CIA_SCOPE_ALL:
        return True
    is_mate = _is_mate_couro_cia(raw_cia)
    if normalized_scope == BI_QUEBRAS_CIA_SCOPE_MATE_COURO:
        return is_mate
    return not is_mate


def _mate_couro_break_totals_date_range(session: Session, d_from: date, d_to: date) -> dict[str, dict[str, int]]:
    """Soma CX/UN de quebras (ChangeLog) no intervalo inclusive, só produtos CIA Mate couro."""
    if d_from > d_to:
        d_from, d_to = d_to, d_from
    if (d_to - d_from).days > 366:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Intervalo maximo 366 dias para acumulado Mate couro.",
        )
    allowed = _mate_couro_allowed_product_codes(session)
    out: dict[str, dict[str, int]] = {}
    cur = d_from
    while cur <= d_to:
        agg = _aggregate_break_events_by_code(session, cur)
        for code, rec in agg.items():
            if code not in allowed:
                continue
            cx = int(rec.get("caixa", 0) or 0)
            un = int(rec.get("unidade", 0) or 0)
            if cx == 0 and un == 0:
                continue
            if code not in out:
                out[code] = {"cx": 0, "un": 0}
            out[code]["cx"] += cx
            out[code]["un"] += un
        cur += timedelta(days=1)
    return {k: v for k, v in out.items() if v["cx"] or v["un"]}


def _mate_couro_break_totals_all_time(session: Session) -> dict[str, dict[str, int]]:
    """Total de quebra CX/UN (servidor) apenas produtos CIA Mate couro; chave numérica canônica."""
    raw = _aggregate_break_events_all_time_by_code(session)
    allowed = _mate_couro_allowed_product_codes(session)
    out: dict[str, dict[str, int]] = {}
    for code, rec in raw.items():
        if code not in allowed:
            continue
        cx = int(rec.get("caixa", 0) or 0)
        un = int(rec.get("unidade", 0) or 0)
        if cx == 0 and un == 0:
            continue
        out[code] = {"cx": cx, "un": un}
    return out


def _merge_break_event_rows_for_operational_day(raw: list[dict]) -> list[dict]:
    """Consolida lançamentos do dia: uma linha por produto com totais líquidos CX e UN no mesmo registro."""
    buckets: dict[str, list[dict]] = defaultdict(list)
    for r in raw:
        code = str(r.get("cod_produto") or "")
        if not code:
            continue
        buckets[code].append(r)

    merged: list[dict] = []
    for code in sorted(buckets.keys()):
        items = buckets[code]
        total_cx = sum(
            int(x.get("quantity") or 0)
            for x in items
            if str(x.get("qty_type") or "caixa") == "caixa"
        )
        total_un = sum(
            int(x.get("quantity") or 0)
            for x in items
            if str(x.get("qty_type") or "caixa") == "unidade"
        )
        if total_cx == 0 and total_un == 0:
            continue
        observed_list = [str(x.get("observed_at") or "") for x in items if x.get("observed_at")]
        observed_max = max(observed_list) if observed_list else None
        actors_set: set[str] = set()
        for x in items:
            a = (x.get("actor") or "").strip()
            if a:
                actors_set.add(a)
        reasons_set: set[str] = set()
        for x in items:
            rs = x.get("reason")
            if rs is not None and str(rs).strip():
                reasons_set.add(str(rs).strip())
        reason_out: str | None = None
        if len(reasons_set) == 1:
            reason_out = next(iter(reasons_set))
        elif len(reasons_set) > 1:
            reason_out = " · ".join(sorted(reasons_set))
        actor_out: str | None = None
        if len(actors_set) == 1:
            actor_out = next(iter(actors_set))
        elif len(actors_set) > 1:
            actor_out = ", ".join(sorted(actors_set))

        merged.append(
            {
                "cod_produto": code,
                "cx": int(total_cx),
                "un": int(total_un),
                "observed_at": observed_max,
                "actor": actor_out,
                "reason": reason_out,
                "client_event_id": None,
                "device_name": None,
            }
        )
    return merged


def _break_event_rows_for_operational_day(session: Session, operational_date: date) -> list[dict]:
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    raw: list[dict] = []
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("operational_date")
        if od:
            try:
                d = date.fromisoformat(str(od)[:10])
            except ValueError:
                continue
            if d != operational_date:
                continue
        else:
            br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
            if br_d is None or br_d != operational_date:
                continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_numeric_product_code_key(item_code)
        if not code:
            continue
        try:
            qty = int(payload.get("quantity", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue
        ct = _extract_count_type(item_code)
        raw.append(
            {
                "cod_produto": code,
                "item_code_raw": item_code,
                "quantity": qty,
                "qty_type": ct,
                "observed_at": payload.get("observed_at"),
                "actor": (log.actor or "").strip() or None,
                "reason": payload.get("reason"),
                "client_event_id": payload.get("client_event_id"),
                "device_name": payload.get("device_name"),
            }
        )

    return _merge_break_event_rows_for_operational_day(raw)


def _break_event_merged_rows_for_code_date_range(
    session: Session, d_from: date, d_to: date, cod_filter: str
) -> list[dict]:
    """Uma passada nos logs: totais de quebra por dia operacional para um único código (chave numérica)."""
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    by_day: dict[date, list[dict]] = defaultdict(list)
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("operational_date")
        if od:
            try:
                d_op = date.fromisoformat(str(od)[:10])
            except ValueError:
                continue
        else:
            br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
            if br_d is None:
                continue
            d_op = br_d
        if d_op < d_from or d_op > d_to:
            continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_numeric_product_code_key(item_code)
        if code != cod_filter:
            continue
        try:
            qty = int(payload.get("quantity", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue
        ct = _extract_count_type(item_code)
        by_day[d_op].append(
            {
                "cod_produto": code,
                "item_code_raw": item_code,
                "quantity": qty,
                "qty_type": ct,
                "observed_at": payload.get("observed_at"),
                "actor": (log.actor or "").strip() or None,
                "reason": payload.get("reason"),
                "client_event_id": payload.get("client_event_id"),
                "device_name": payload.get("device_name"),
            }
        )

    out: list[dict] = []
    for d_op in sorted(by_day.keys()):
        merged = _merge_break_event_rows_for_operational_day(by_day[d_op])
        for r in merged:
            rr = dict(r)
            rr["operational_date"] = d_op.isoformat()
            out.append(rr)
    return out


@router.get("/break-day-totals")
def break_day_totals(
    operational_date: str | None = Query(
        default=None,
        description="YYYY-MM-DD do dia operacional (America/Sao_Paulo). Padrão: hoje.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    d = _parse_iso_date_arg(operational_date) if operational_date else br_today
    if d is None:
        d = br_today
    agg = _aggregate_break_events_by_code(session, d)
    balances: dict[str, dict[str, int]] = {}
    for code, rec in agg.items():
        balances[code] = {
            "caixa": int(rec.get("caixa", 0)),
            "unidade": int(rec.get("unidade", 0)),
        }
    return {"operational_date": d.isoformat(), "balances": balances}


@router.get("/break-events")
def list_break_events(
    operational_date: str | None = Query(
        default=None,
        description="YYYY-MM-DD do dia operacional (America/Sao_Paulo). Padrão: hoje.",
    ),
    date_from: str | None = Query(
        default=None,
        description="Com date_to e cod_produto: início do intervalo (YYYY-MM-DD), inclusive.",
    ),
    date_to: str | None = Query(
        default=None,
        description="Com date_from e cod_produto: fim do intervalo (YYYY-MM-DD), inclusive.",
    ),
    cod_produto: str | None = Query(
        default=None,
        description="Com date_from e date_to: filtra um produto (chave numérica canônica).",
    ),
    limit: int = Query(default=2000, ge=1, le=5000),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    df_r = _parse_iso_date_arg(date_from) if date_from else None
    dt_r = _parse_iso_date_arg(date_to) if date_to else None
    cod_q = (cod_produto or "").strip()

    if (df_r is not None) ^ (dt_r is not None):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Informe date_from e date_to juntos (YYYY-MM-DD).",
        )
    if df_r is not None and dt_r is not None:
        if not cod_q:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Consulta por intervalo exige cod_produto.",
            )
        d0, d1 = df_r, dt_r
        if d0 > d1:
            d0, d1 = d1, d0
        if (d1 - d0).days > 120:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Intervalo maximo 120 dias para quebra por produto.",
            )
        c = _normalize_numeric_product_code_key(cod_q)
        if not c:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="cod_produto invalido.",
            )
        rows = _break_event_merged_rows_for_code_date_range(session, d0, d1, c)[:limit]
        desc_map: dict[str, str] = {}
        codes_lookup: set[str] = {c}
        alt_cod = _normalize_item_code(cod_q)
        if alt_cod:
            codes_lookup.add(alt_cod)
        prod_rows = list(
            session.exec(select(Product).where(Product.cod_produto.in_(list(codes_lookup)))).all()
        )
        for p in prod_rows:
            cc = _normalize_numeric_product_code_key(str(p.cod_produto or ""))
            if cc:
                desc_map[cc] = (p.cod_grup_descricao or "").strip()
        break_logins: set[str] = set()
        for r in rows:
            ac = r.get("actor")
            if ac and str(ac).strip():
                break_logins.update(p.strip() for p in re.split(r",\s*", str(ac).strip()) if p.strip())
        break_name_map = _display_name_map_for_logins(session, break_logins)
        for r in rows:
            rc = str(r.get("cod_produto") or "")
            r["product_desc"] = desc_map.get(rc) or None
            ac = r.get("actor")
            if ac:
                disp = _actor_csv_to_display_labels(str(ac), break_name_map)
                r["actor"] = disp if disp else ac
        return {
            "date_from": d0.isoformat(),
            "date_to": d1.isoformat(),
            "operational_date": None,
            "cod_produto": c,
            "events": rows,
            "count": len(rows),
        }

    d = _parse_iso_date_arg(operational_date) if operational_date else br_today
    if d is None:
        d = br_today
    rows = _break_event_rows_for_operational_day(session, d)[:limit]
    codes = list({str(r.get("cod_produto") or "") for r in rows if r.get("cod_produto")})
    desc_map: dict[str, str] = {}
    if codes:
        prod_rows = list(session.exec(select(Product).where(Product.cod_produto.in_(codes))).all())
        for p in prod_rows:
            c = (p.cod_produto or "").strip()
            if c:
                desc_map[c] = (p.cod_grup_descricao or "").strip()
    break_logins: set[str] = set()
    for r in rows:
        ac = r.get("actor")
        if ac and str(ac).strip():
            break_logins.update(p.strip() for p in re.split(r",\s*", str(ac).strip()) if p.strip())
    break_name_map = _display_name_map_for_logins(session, break_logins)
    for r in rows:
        c = str(r.get("cod_produto") or "")
        r["product_desc"] = desc_map.get(c) or None
        ac = r.get("actor")
        if ac:
            disp = _actor_csv_to_display_labels(str(ac), break_name_map)
            r["actor"] = disp if disp else ac
    return {"operational_date": d.isoformat(), "events": rows, "count": len(rows)}


@router.post("/break-events")
def ingest_break_events(
    payload: BreakEventsPayload,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    synced_ids: list[str] = []
    max_pg_int = 2_147_483_647

    try:
        for event in payload.events:
            if event.quantity == 0:
                continue
            dt_utc, br_date = _parse_and_validate_observed_at(event.observed_at)
            obs_stored = dt_utc.replace(microsecond=0).isoformat().replace("+00:00", "Z")

            op_d: date | None = None
            if event.operational_date:
                op_d = _parse_iso_date_arg(event.operational_date)
            if op_d is None:
                op_d = br_date

            event_hash = hashlib.sha256(event.client_event_id.encode("utf-8")).hexdigest()
            entity_id = (int(event_hash[:16], 16) % max_pg_int) + 1

            existing = session.exec(
                select(ChangeLog).where(
                    ChangeLog.entity_name == "stock_break",
                    ChangeLog.action == "break_event",
                    ChangeLog.entity_id == entity_id,
                )
            ).first()
            if existing:
                synced_ids.append(event.client_event_id)
                continue

            log = ChangeLog(
                entity_name="stock_break",
                entity_id=entity_id,
                action="break_event",
                actor=user.username,
                changed_at=datetime.now(timezone.utc),
                payload={
                    "client_event_id": event.client_event_id,
                    "item_code": event.item_code,
                    "quantity": event.quantity,
                    "observed_at": obs_stored,
                    "device_name": event.device_name,
                    "reason": event.reason,
                    "operational_date": op_d.isoformat(),
                },
            )
            session.add(log)
            synced_ids.append(event.client_event_id)

        session.commit()
        return {"received": len(payload.events), "synced": len(synced_ids), "synced_ids": synced_ids}
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Falha ao sincronizar eventos de quebra")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Falha ao sincronizar quebra: {exc}",
        )


@router.post("/break-events/bulk-delete")
def bulk_delete_break_events(
    body: BreakEventsBulkDeleteBody,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("administrativo", "admin")),
) -> dict:
    """Apaga em lote eventos ``stock_break`` / ``break_event`` do dia operacional (servidor).

    - Sem ``cod_produtos``: remove todas as quebras daquele dia; exige frase exata de confirmação.
    - Com ``cod_produtos``: remove só lançamentos cujo código (canônico numérico) está na lista.
    """
    op_d = _parse_iso_date_arg(body.operational_date)
    if op_d is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="operational_date invalido (use YYYY-MM-DD).",
        )

    raw_codes = [str(x).strip() for x in (body.cod_produtos or []) if str(x).strip()]
    cod_filters: set[str] = set()
    for raw in raw_codes:
        c = _normalize_numeric_product_code_key(raw)
        if c:
            cod_filters.add(c)

    phrase = (body.confirm_phrase or "").strip()
    scoped_by_code = len(raw_codes) > 0
    if scoped_by_code:
        if not cod_filters:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Nenhum codigo valido em cod_produtos.",
            )
        targets = _collect_break_event_logs_for_day_and_codes(session, op_d, cod_filters)
    else:
        if phrase != BREAK_BULK_DELETE_DAY_PHRASE:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=(
                    "Para apagar todas as quebras do dia, confirme com a frase exata: "
                    f"{BREAK_BULK_DELETE_DAY_PHRASE}"
                ),
            )
        targets = _collect_break_event_logs_for_day_and_codes(session, op_d, None)

    if not targets:
        return {
            "deleted": 0,
            "operational_date": op_d.isoformat(),
            "cod_produtos": sorted(cod_filters) if cod_filters else None,
            "message": "Nenhum lançamento encontrado para os critérios.",
        }

    deleted_ids: list[int] = []
    try:
        for log in targets:
            if log.id is not None:
                deleted_ids.append(int(log.id))
            session.delete(log)

        audit = ChangeLog(
            entity_name="stock_break",
            entity_id=-1,
            action="break_bulk_delete",
            actor=user.username,
            changed_at=datetime.now(timezone.utc),
            payload={
                "operational_date": op_d.isoformat(),
                "deleted_count": len(targets),
                "deleted_change_log_ids": deleted_ids[:500],
                "cod_produtos_filter": sorted(cod_filters) if cod_filters else None,
            },
        )
        session.add(audit)
        session.commit()
        return {
            "deleted": len(targets),
            "operational_date": op_d.isoformat(),
            "cod_produtos": sorted(cod_filters) if cod_filters else None,
            "message": f"Removidos {len(targets)} lançamento(s) de quebra no servidor.",
        }
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Falha ao apagar quebras em lote")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Falha ao apagar quebras: {exc}",
        ) from exc


@router.get("/validity-lines/{line_id}")
def get_validity_line(
    line_id: int,
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Uma linha por id (evita 405 em clientes que fazem GET no mesmo path do DELETE)."""
    _ensure_validity_lines_table()
    row = session.get(ValidityLine, line_id)
    if not row:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Linha de validade nao encontrada")
    try:
        v_logins = (
            {(row.actor_username or "").strip()}
            if row.actor_username and str(row.actor_username).strip()
            else set()
        )
        v_name_map = _display_name_map_for_logins(session, v_logins)
        au = row.actor_username
        au_disp = (
            v_name_map.get((au or "").strip(), au)
            if au and str(au).strip()
            else au
        )
        return {
            "id": row.id,
            "client_event_id": row.client_event_id,
            "cod_produto": row.cod_produto,
            "expiration_date": row.expiration_date.isoformat(),
            "quantity_un": int(row.quantity_un),
            "quantity_cx": int(row.quantity_cx or 0),
            "lot_code": row.lot_code,
            "note": row.note,
            "operational_date": row.operational_date.isoformat(),
            "observed_at": _validity_observed_at_iso(row.observed_at),
            "device_name": row.device_name,
            "actor_username": au_disp,
        }
    except SQLAlchemyError as exc:
        logger.exception("Erro ao ler linha de validade %s", line_id)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao consultar validade: {exc}",
        ) from exc


@router.get("/validity-lines")
def list_validity_lines(
    operational_date: str | None = Query(
        default=None,
        description="YYYY-MM-DD do dia operacional (America/Sao_Paulo). Padrão: hoje.",
    ),
    include_all_days: bool = Query(
        default=False,
        description="Quando true, retorna linhas de todas as datas operacionais.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    _ensure_validity_lines_table()
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    d = _parse_iso_date_arg(operational_date) if operational_date else br_today
    if d is None:
        d = br_today
    try:
        stmt = select(ValidityLine)
        if not include_all_days:
            stmt = stmt.where(ValidityLine.operational_date == d)
        rows = list(
            session.exec(
                stmt.order_by(
                    ValidityLine.cod_produto,
                    ValidityLine.operational_date,
                    ValidityLine.expiration_date,
                    ValidityLine.id,
                )
            ).all()
        )
    except SQLAlchemyError as exc:
        logger.exception("Erro ao listar validity_lines")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao consultar validades: {exc}",
        ) from exc
    v_logins = {
        (r.actor_username or "").strip()
        for r in rows
        if r.actor_username and str(r.actor_username).strip()
    }
    v_name_map = _display_name_map_for_logins(session, v_logins)
    lines = []
    for r in rows:
        au = r.actor_username
        au_disp = (
            v_name_map.get((au or "").strip(), au)
            if au and str(au).strip()
            else au
        )
        lines.append(
            {
                "id": r.id,
                "client_event_id": r.client_event_id,
                "cod_produto": r.cod_produto,
                "expiration_date": r.expiration_date.isoformat(),
                "quantity_un": int(r.quantity_un),
                "quantity_cx": int(r.quantity_cx or 0),
                "lot_code": r.lot_code,
                "note": r.note,
                "operational_date": r.operational_date.isoformat(),
                "observed_at": _validity_observed_at_iso(r.observed_at),
                "device_name": r.device_name,
                "actor_username": au_disp,
            }
        )
    return {"operational_date": d.isoformat(), "lines": lines}


@router.get("/validity-last-launch-by-product")
def validity_last_launch_by_product(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Última data operacional (America/Sao_Paulo) em que cada produto teve ao menos uma linha de validade."""
    _ensure_validity_lines_table()
    try:
        rows = session.exec(
            select(ValidityLine.cod_produto, func.max(ValidityLine.operational_date)).group_by(
                ValidityLine.cod_produto
            )
        ).all()
    except SQLAlchemyError as exc:
        logger.exception("Erro ao agregar ultima validade por produto")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao consultar ultimas validades: {exc}",
        ) from exc
    last_by_code: dict[str, str] = {}
    for cod_raw, d in rows:
        if d is None:
            continue
        cod = _normalize_item_code(cod_raw)
        if not cod:
            continue
        last_by_code[cod] = d.isoformat()
    return {"last_by_code": last_by_code}


def _build_validity_display_expiry_by_code(session: Session, br_today: date) -> dict[str, str]:
    """Data de exibição por código: primeiro vencimento ≥ hoje; se só há vencidos, o mais antigo."""
    rows = session.exec(select(ValidityLine.cod_produto, ValidityLine.expiration_date)).all()
    by_code: dict[str, list[date]] = defaultdict(list)
    for cod_raw, exp in rows:
        if exp is None:
            continue
        cod = _normalize_item_code(cod_raw)
        if not cod:
            continue
        by_code[cod].append(exp)
    out: dict[str, str] = {}
    for cod, dates in by_code.items():
        uniq = sorted(set(dates))
        chosen: date | None = None
        for d in uniq:
            if d >= br_today:
                chosen = d
                break
        if chosen is None and uniq:
            chosen = uniq[0]
        if chosen is not None:
            out[cod] = chosen.isoformat()
    return out


@router.get("/validity-display-expiry-by-product")
def validity_display_expiry_by_product(
    session: Session = Depends(get_session),
    _: User = Depends(require_stock_analysis_access),
) -> dict:
    """
    Mapa código → data de vencimento para exibição (alinhado ao módulo Validade no front:
    primeiro vencimento hoje ou futuro; se só há linhas vencidas, a mais antiga).
    """
    _ensure_validity_lines_table()
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    try:
        out = _build_validity_display_expiry_by_code(session, br_today)
    except SQLAlchemyError as exc:
        logger.exception("Erro ao agregar validade por produto")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao consultar validades agregadas: {exc}",
        ) from exc
    return {"today": br_today.isoformat(), "by_code": out}


@router.get("/validity-kpi-expiring-30d")
def validity_kpi_expiring_30d(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """
    Quantidade de produtos ativos cuja data de validade exibida (mesma regra do módulo Validade)
    está entre hoje e hoje+30 dias (inclusive). Não inclui já vencidos.
    """
    from app.api.routes.products import _catalog_status_is_ativo_clause

    _ensure_validity_lines_table()
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    try:
        by_code = _build_validity_display_expiry_by_code(session, br_today)
        active_rows = session.exec(select(Product.cod_produto).where(_catalog_status_is_ativo_clause())).all()
        active_set = {_normalize_item_code(c) for c in active_rows if c}
    except SQLAlchemyError as exc:
        logger.exception("Erro ao calcular KPI de validade 30d")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao consultar validade para KPI: {exc}",
        ) from exc
    n = 0
    for cod, iso in by_code.items():
        if cod not in active_set:
            continue
        try:
            chosen = date.fromisoformat(iso[:10])
        except ValueError:
            continue
        delta = (chosen - br_today).days
        if 0 <= delta <= 30:
            n += 1
    return {"today": br_today.isoformat(), "window_days": 30, "count": n}


@router.delete("/validity-lines/{line_id}", status_code=204)
def delete_validity_line(
    line_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> None:
    _ensure_validity_lines_table()
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    row = session.get(ValidityLine, line_id)
    if not row:
        raise HTTPException(status_code=404, detail="Linha de validade nao encontrada")
    if row.operational_date != br_today:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="So e permitido excluir lancamentos do dia corrente (America/Sao_Paulo).",
        )
    session.delete(row)
    session.commit()


@router.post("/validity-events")
def ingest_validity_events(
    payload: ValidityEventsPayload,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Sincroniza lançamentos de validade (idempotente por client_event_id)."""
    _ensure_validity_lines_table()
    synced_ids: list[str] = []
    now_br = datetime.now(timezone.utc).astimezone(_BR)
    br_today = now_br.date()

    ref_d = _parse_iso_date_arg(payload.reference_date) if payload.reference_date else None
    un_map = _un_balance_map_for_validity(session, ref_d)
    cx_map = _cx_balance_map_for_validity(session, ref_d)
    count_day = ref_d if ref_d is not None else br_today
    counted_by_code = _aggregate_count_events_by_code(session, count_on_date=count_day)

    def _validity_un_cap(code: str) -> int:
        c = _normalize_item_code(code)
        if not c:
            return 0
        if c in counted_by_code:
            return int(counted_by_code[c].get("unidade", 0))
        return int(un_map.get(c, 0))

    def _validity_cx_cap(code: str) -> int:
        c = _normalize_item_code(code)
        if not c:
            return 0
        if c in counted_by_code:
            return int(counted_by_code[c].get("caixa", 0))
        return int(cx_map.get(c, 0))

    # Quantidades já gravadas hoje por código (excluindo duplicatas de idempotência)
    existing_by_un: dict[str, int] = {}
    existing_by_cx: dict[str, int] = {}
    existing_rows = list(
        session.exec(select(ValidityLine).where(ValidityLine.operational_date == br_today)).all()
    )
    for er in existing_rows:
        c = _normalize_item_code(er.cod_produto)
        if not c:
            continue
        existing_by_un[c] = existing_by_un.get(c, 0) + int(er.quantity_un)
        existing_by_cx[c] = existing_by_cx.get(c, 0) + int(er.quantity_cx or 0)

    # Novas quantidades por código neste batch (apenas eventos que virarão insert)
    new_by_un: dict[str, int] = {}
    new_by_cx: dict[str, int] = {}
    events_to_insert: list[ValidityEventInput] = []

    try:
        for event in payload.events:
            if event.quantity_un < 0 or event.quantity_cx < 0:
                continue
            dt_utc, br_date = _parse_and_validate_observed_at(event.observed_at)
            if br_date != br_today:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Lancamento de validade apenas para o dia corrente (America/Sao_Paulo).",
                )
            code = _normalize_item_code(event.cod_produto)
            if not code:
                continue

            dup = session.exec(
                select(ValidityLine).where(ValidityLine.client_event_id == event.client_event_id)
            ).first()
            if dup:
                synced_ids.append(event.client_event_id)
                continue

            events_to_insert.append(event)
            new_by_un[code] = new_by_un.get(code, 0) + int(event.quantity_un)
            new_by_cx[code] = new_by_cx.get(code, 0) + int(event.quantity_cx)

        # Teto UN: só valida se houver lançamento com UN neste batch (compatível com fluxo legado).
        touched = set(new_by_un) | set(new_by_cx)
        for code in touched:
            add_un = new_by_un.get(code, 0)
            if add_un > 0:
                cap = _validity_un_cap(code)
                total_after = existing_by_un.get(code, 0) + add_un
                if cap <= 0 and total_after > 0:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Produto {code}: base UN zero (contagem/TXT); nao e possivel classificar quantidade.",
                    )
                if cap > 0 and total_after > cap:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            f"Produto {code}: soma das quantidades por validade ({total_after} UN) "
                            f"excede a base ({cap} UN) da contagem ou do TXT."
                        ),
                    )
            add_cx = new_by_cx.get(code, 0)
            if add_cx > 0:
                cap_cx = _validity_cx_cap(code)
                total_cx = existing_by_cx.get(code, 0) + add_cx
                if cap_cx <= 0 and total_cx > 0:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Produto {code}: base CX zero (contagem/TXT); nao e possivel classificar caixas.",
                    )
                if cap_cx > 0 and total_cx > cap_cx:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=(
                            f"Produto {code}: soma das caixas por validade ({total_cx} CX) "
                            f"excede a base ({cap_cx} CX) da contagem ou do TXT."
                        ),
                    )

        for event in events_to_insert:
            dt_utc, br_date = _parse_and_validate_observed_at(event.observed_at)
            code = _normalize_item_code(event.cod_produto)
            row = ValidityLine(
                client_event_id=event.client_event_id,
                cod_produto=code,
                expiration_date=event.expiration_date,
                quantity_un=int(event.quantity_un),
                quantity_cx=int(event.quantity_cx),
                lot_code=(event.lot_code or "").strip() or None,
                note=(event.note or "").strip() or None,
                operational_date=br_date,
                observed_at=dt_utc.replace(tzinfo=timezone.utc),
                device_name=event.device_name,
                actor_username=user.username,
            )
            session.add(row)
            synced_ids.append(event.client_event_id)

        session.commit()
        return {"received": len(payload.events), "synced": len(synced_ids), "synced_ids": synced_ids}
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Falha ao sincronizar validade")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Falha ao sincronizar validade: {exc}",
        ) from exc


def _ensure_recount_signals_table() -> None:
    try:
        SQLModel.metadata.create_all(engine, tables=[RecountSignal.__table__], checkfirst=True)
    except SQLAlchemyError:
        logger.exception("Falha ao garantir tabela recount_signals")
        raise


def _brazil_today_date() -> date:
    return datetime.now(timezone.utc).astimezone(_BR).date()


class RecountSignalIn(BaseModel):
    cod_produto: str = Field(min_length=1, max_length=120)
    operational_date: date | None = None


class RecountSignalsOut(BaseModel):
    operational_date: date
    codes: list[str]


@router.post("/recount-signal")
def post_recount_signal(
    body: RecountSignalIn,
    session: Session = Depends(get_session),
    user: User = Depends(require_stock_analysis_access),
):
    """Analista solicita recontagem ao conferente para o dia operacional informado (ou hoje em SP)."""
    _ensure_recount_signals_table()
    cod = _normalize_item_code(body.cod_produto)
    if not cod:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Codigo do produto invalido.")
    op_date = body.operational_date or _brazil_today_date()
    existing = session.exec(
        select(RecountSignal).where(
            RecountSignal.operational_date == op_date,
            RecountSignal.cod_produto == cod,
        )
    ).first()
    now = datetime.now(timezone.utc)
    if existing:
        existing.requested_at = now
        existing.requested_by = user.username
        session.add(existing)
    else:
        session.add(
            RecountSignal(
                operational_date=op_date,
                cod_produto=cod,
                requested_by=user.username,
                requested_at=now,
            )
        )
    session.commit()
    return {"ok": True, "cod_produto": cod, "operational_date": op_date.isoformat()}


@router.get("/recount-signals", response_model=RecountSignalsOut)
def get_recount_signals(
    operational_date: date | None = Query(default=None),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
):
    """Lista códigos com solicitação de recontagem ativa para o dia operacional (alinhado à #count-date)."""
    _ensure_recount_signals_table()
    op_date = operational_date or _brazil_today_date()
    rows = session.exec(select(RecountSignal).where(RecountSignal.operational_date == op_date)).all()
    codes = sorted({str(r.cod_produto) for r in rows})
    return RecountSignalsOut(operational_date=op_date, codes=codes)


MATE_COURO_CIA = "Mate couro"
_ALLOWED_MATE_TROCA_KINDS = frozenset(
    {"chegada", "definir", "zerar", "ajuste_pendente", "incorporacao_quebra"}
)
MAX_MATE_TROCA_BATCH_SCAN = 12_000
BI_QUEBRAS_CIA_SCOPE_ALL = "todas"
BI_QUEBRAS_CIA_SCOPE_MATE_COURO = "mate-couro"
BI_QUEBRAS_CIA_SCOPE_OUTRAS = "outras"


def _ensure_mate_couro_troca_logs_table() -> None:
    try:
        SQLModel.metadata.create_all(engine, tables=[MateCouroTrocaLog.__table__], checkfirst=True)
    except SQLAlchemyError:
        logger.exception("Falha ao garantir tabela mate_couro_troca_logs")
        raise


def _mate_troca_normalize_cx_un(cx: int, un: int, factor: float | None) -> tuple[int, int]:
    """Converte UN acumuladas em CX inteiras quando UN >= fator (UN por 1 CX). Só aplica com fator inteiro > 0."""
    cx_i = max(0, int(cx))
    un_i = max(0, int(un))
    if factor is None:
        return cx_i, un_i
    try:
        f = float(factor)
    except (TypeError, ValueError):
        return cx_i, un_i
    if not (f > 0) or f != f:  # NaN
        return cx_i, un_i
    fr = round(f)
    if fr <= 0 or abs(f - fr) > 1e-9:
        return cx_i, un_i
    fi = int(fr)
    total = cx_i * fi + un_i
    return total // fi, total % fi


def _mate_troca_product_factor_by_code(session: Session) -> dict[str, float]:
    """Mapa código canônico (numérico) → fator de conversão; só produtos CIA Mate couro com fator definido."""
    out: dict[str, float] = {}
    rows = list(
        session.exec(
            select(Product.cod_produto, Product.conversion_factor).where(Product.cod_grup_cia == MATE_COURO_CIA)
        ).all()
    )
    for cod_produto, cf in rows:
        if cf is None:
            continue
        try:
            fv = float(cf)
        except (TypeError, ValueError):
            continue
        ck = _normalize_numeric_product_code_key(str(cod_produto or ""))
        if not ck:
            continue
        out[ck] = fv
    return out


class MateTrocaEventInput(BaseModel):
    client_event_id: str = Field(min_length=8, max_length=100)
    kind: str = Field(max_length=24)
    cod_produto: str = Field(min_length=1, max_length=120)
    qty_cx_in: int = Field(default=0, ge=-500_000, le=500_000)
    qty_un_in: int = Field(default=0, ge=-500_000, le=500_000)
    pend_cx_before: int = Field(ge=0, le=500_000)
    pend_un_before: int = Field(ge=0, le=500_000)
    pend_cx_after: int = Field(ge=0, le=500_000)
    pend_un_after: int = Field(ge=0, le=500_000)
    excess_cx: int = Field(default=0, ge=0, le=500_000)
    excess_un: int = Field(default=0, ge=0, le=500_000)
    device_name: str | None = Field(default=None, max_length=120)


class MateTrocaEventsPayload(BaseModel):
    events: list[MateTrocaEventInput]


def _canonical_mate_couro_cod(session: Session, raw: str) -> str:
    c = _normalize_item_code(raw)
    if not c:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Codigo invalido.",
        )
    prod = session.exec(
        select(Product).where(Product.cod_grup_cia == MATE_COURO_CIA, Product.cod_produto == c)
    ).first()
    if prod:
        return _normalize_item_code(prod.cod_produto or c)
    if c.isdigit():
        alt = str(int(c))
        prod2 = session.exec(
            select(Product).where(Product.cod_grup_cia == MATE_COURO_CIA, Product.cod_produto == alt)
        ).first()
        if prod2:
            return _normalize_item_code(prod2.cod_produto or alt)
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail=f"Produto nao encontrado ou CIA diferente de {MATE_COURO_CIA}.",
    )


def _validate_mate_troca_payload(ev: MateTrocaEventInput, factor: float | None) -> tuple[int, int]:
    """Retorna (excess_cx, excess_un) gravados; pendente após evento é normalizado pelo fator (UN por CX) quando aplicável."""
    k = (ev.kind or "").strip()
    if k == "chegada":
        if ev.qty_cx_in < 0 or ev.qty_un_in < 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="chegada exige quantidades nao negativas.",
            )
        nbcx, nbun = _mate_troca_normalize_cx_un(ev.pend_cx_before, ev.pend_un_before, factor)
        ex_cx = max(0, ev.qty_cx_in - nbcx)
        ex_un = max(0, ev.qty_un_in - nbun)
        raw_acx = max(0, nbcx - ev.qty_cx_in)
        raw_aun = max(0, nbun - ev.qty_un_in)
        exp_cx, exp_un = _mate_troca_normalize_cx_un(raw_acx, raw_aun, factor)
        if ev.pend_cx_after != exp_cx or ev.pend_un_after != exp_un:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="pendente apos chegada inconsistente (verifique fator de conversao e saldo atual).",
            )
        return ex_cx, ex_un
    if k in ("definir", "ajuste_pendente"):
        if ev.qty_cx_in < 0 or ev.qty_un_in < 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="definir/ajuste exige quantidades nao negativas.",
            )
        exp_cx, exp_un = _mate_troca_normalize_cx_un(ev.qty_cx_in, ev.qty_un_in, factor)
        if ev.pend_cx_after != exp_cx or ev.pend_un_after != exp_un:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Pendente apos ajuste deve coincidir com totais normalizados (CX/UN e fator).",
            )
        return 0, 0
    if k == "zerar":
        if ev.qty_cx_in != 0 or ev.qty_un_in != 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="zerar exige quantidades de entrada zero.",
            )
        if ev.pend_cx_after != 0 or ev.pend_un_after != 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="zerar exige pendente apos zero.",
            )
        return 0, 0
    if k == "incorporacao_quebra":
        nbcx, nbun = _mate_troca_normalize_cx_un(ev.pend_cx_before, ev.pend_un_before, factor)
        raw_cx = max(0, nbcx + int(ev.qty_cx_in))
        raw_un = max(0, nbun + int(ev.qty_un_in))
        exp_cx, exp_un = _mate_troca_normalize_cx_un(raw_cx, raw_un, factor)
        if int(ev.pend_cx_after) != exp_cx or int(ev.pend_un_after) != exp_un:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Pendente apos incorporacao_quebra inconsistente com delta (inclui conversao CX/UN).",
            )
        return 0, 0
    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="kind invalido.")


@router.post("/mate-troca-events")
def ingest_mate_troca_events(
    payload: MateTrocaEventsPayload,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Sincroniza log de chegadas/ajustes da Base de Troca (idempotente por client_event_id)."""
    _ensure_mate_couro_troca_logs_table()
    synced_ids: list[str] = []

    try:
        fac_map = _mate_troca_product_factor_by_code(session)
        for ev in payload.events:
            if (ev.kind or "").strip() not in _ALLOWED_MATE_TROCA_KINDS:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"kind invalido: {ev.kind}",
                )
            dup = session.exec(
                select(MateCouroTrocaLog).where(MateCouroTrocaLog.client_event_id == ev.client_event_id)
            ).first()
            if dup:
                synced_ids.append(ev.client_event_id)
                continue

            cod = _canonical_mate_couro_cod(session, ev.cod_produto)
            fac = fac_map.get(cod)
            ex_cx, ex_un = _validate_mate_troca_payload(ev, fac)

            row = MateCouroTrocaLog(
                client_event_id=ev.client_event_id,
                kind=(ev.kind or "").strip(),
                cod_produto=cod,
                qty_cx_in=int(ev.qty_cx_in),
                qty_un_in=int(ev.qty_un_in),
                pend_cx_before=int(ev.pend_cx_before),
                pend_un_before=int(ev.pend_un_before),
                pend_cx_after=int(ev.pend_cx_after),
                pend_un_after=int(ev.pend_un_after),
                excess_cx=int(ex_cx),
                excess_un=int(ex_un),
                device_name=(ev.device_name or "").strip() or None,
                actor_username=user.username,
                created_at=datetime.now(timezone.utc),
            )
            session.add(row)
            synced_ids.append(ev.client_event_id)

        session.commit()
        return {"received": len(payload.events), "synced": len(synced_ids), "synced_ids": synced_ids}
    except HTTPException:
        session.rollback()
        raise
    except Exception as exc:
        session.rollback()
        logger.exception("Falha ao sincronizar mate-troca-events")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Falha ao sincronizar base de troca: {exc}",
        ) from exc


@router.get("/mate-troca-events")
def list_mate_troca_events(
    date_from: str | None = Query(default=None, description="YYYY-MM-DD filtro em America/Sao_Paulo"),
    date_to: str | None = Query(default=None, description="YYYY-MM-DD inclusive em America/Sao_Paulo"),
    cod_produto: str | None = Query(default=None),
    limit: int = Query(default=200, ge=1, le=500),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Lista auditoria da Base de Troca (Mate couro), mais recentes primeiro."""
    _ensure_mate_couro_troca_logs_table()
    stmt = select(MateCouroTrocaLog).order_by(MateCouroTrocaLog.created_at.desc())

    df = _parse_iso_date_arg(date_from) if date_from else None
    dt = _parse_iso_date_arg(date_to) if date_to else None
    if df is not None:
        start_utc = datetime.combine(df, time.min, tzinfo=_BR).astimezone(timezone.utc)
        stmt = stmt.where(MateCouroTrocaLog.created_at >= start_utc.replace(tzinfo=None))
    if dt is not None:
        end_br = dt + timedelta(days=1)
        end_utc = datetime.combine(end_br, time.min, tzinfo=_BR).astimezone(timezone.utc)
        stmt = stmt.where(MateCouroTrocaLog.created_at < end_utc.replace(tzinfo=None))

    if (cod_produto or "").strip():
        c = _normalize_item_code(cod_produto)
        stmt = stmt.where(MateCouroTrocaLog.cod_produto == c)

    rows = list(session.exec(stmt.limit(limit)).all())
    codes = list({str(r.cod_produto or "") for r in rows if r.cod_produto})
    desc_map: dict[str, str] = {}
    if codes:
        prods = list(session.exec(select(Product).where(Product.cod_produto.in_(codes))).all())
        for p in prods:
            cc = _normalize_item_code(p.cod_produto or "")
            if cc:
                desc_map[cc] = (p.cod_grup_descricao or "").strip()

    mt_logins = {
        (r.actor_username or "").strip()
        for r in rows
        if r.actor_username and str(r.actor_username).strip()
    }
    mt_name_map = _display_name_map_for_logins(session, mt_logins)

    def _iso(dt_val: datetime | None) -> str | None:
        if dt_val is None:
            return None
        u = dt_val if dt_val.tzinfo else dt_val.replace(tzinfo=timezone.utc)
        return u.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    events = []
    for r in rows:
        c = str(r.cod_produto or "")
        events.append(
            {
                "id": r.id,
                "client_event_id": r.client_event_id,
                "kind": r.kind,
                "cod_produto": c,
                "product_desc": desc_map.get(c) or None,
                "qty_cx_in": int(r.qty_cx_in),
                "qty_un_in": int(r.qty_un_in),
                "pend_cx_before": int(r.pend_cx_before),
                "pend_un_before": int(r.pend_un_before),
                "pend_cx_after": int(r.pend_cx_after),
                "pend_un_after": int(r.pend_un_after),
                "excess_cx": int(r.excess_cx),
                "excess_un": int(r.excess_un),
                "device_name": r.device_name,
                "actor_username": (
                    mt_name_map.get((r.actor_username or "").strip(), r.actor_username)
                    if r.actor_username
                    else None
                ),
                "created_at": _iso(r.created_at),
            }
        )

    return {"count": len(events), "events": events}


def _mate_troca_pending_product_map(session: Session) -> dict[str, dict[str, int]]:
    """Pendente CX/UN por código canônico (último evento por produto, aliases numéricos unificados).

    Saldo é normalizado pelo fator de conversão do cadastro (UN por CX), quando o fator é inteiro > 0.
    """
    _ensure_mate_couro_troca_logs_table()
    fac_map = _mate_troca_product_factor_by_code(session)
    rn = (
        func.row_number()
        .over(
            partition_by=MateCouroTrocaLog.cod_produto,
            order_by=(desc(MateCouroTrocaLog.created_at), desc(MateCouroTrocaLog.id)),
        )
        .label("rn")
    )
    sub = select(
        MateCouroTrocaLog.cod_produto,
        MateCouroTrocaLog.pend_cx_after,
        MateCouroTrocaLog.pend_un_after,
        MateCouroTrocaLog.created_at,
        MateCouroTrocaLog.id,
        rn,
    ).subquery()
    stmt = select(
        sub.c.cod_produto,
        sub.c.pend_cx_after,
        sub.c.pend_un_after,
        sub.c.created_at,
        sub.c.id,
    ).where(sub.c.rn == 1)
    rows = list(session.exec(stmt).all())
    pending: dict[str, dict[str, int]] = {}
    best_row: dict[str, tuple[int, int, datetime | None, int | None, str]] = {}
    for row in rows:
        cod_raw, pcx, pun, cr_at, rid = row[0], row[1], row[2], row[3], row[4]
        cx = int(pcx or 0)
        un = int(pun or 0)
        if cx == 0 and un == 0:
            continue
        cod_raw_str = str(cod_raw or "")
        c = _normalize_numeric_product_code_key(cod_raw_str)
        if not c:
            continue
        cur = best_row.get(c)
        cand = (cx, un, cr_at, rid, cod_raw_str)
        if cur is None:
            best_row[c] = cand
            continue
        cand_pref = _mate_troca_cod_preferred_over_alias(cod_raw_str, c)
        cur_pref = _mate_troca_cod_preferred_over_alias(cur[4], c)
        if cand_pref and not cur_pref:
            best_row[c] = cand
            continue
        if cur_pref and not cand_pref:
            continue
        _, _, cr_at, rid, _ = cand
        _, _, cur_at, cur_id, _ = cur
        newer = False
        if cr_at is not None and cur_at is not None:
            newer = cr_at > cur_at or (cr_at == cur_at and (rid or 0) > (cur_id or 0))
        elif cr_at is not None and cur_at is None:
            newer = True
        elif cr_at is None and cur_at is None:
            newer = (rid or 0) > (cur_id or 0)
        if newer:
            best_row[c] = cand
    for c, (cx, un, _, _, _) in best_row.items():
        if cx == 0 and un == 0:
            continue
        fac = fac_map.get(c)
        ncx, nun = _mate_troca_normalize_cx_un(cx, un, fac)
        pending[c] = {"cx": ncx, "un": nun}
    return pending


def _mate_troca_created_br_date(dt: datetime | None) -> date | None:
    if dt is None:
        return None
    u = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    return u.astimezone(_BR).date()


def _mate_troca_events_in_period_br(evs: list[MateCouroTrocaLog], d0: date, d1: date) -> list[MateCouroTrocaLog]:
    out: list[MateCouroTrocaLog] = []
    for e in evs:
        bd = _mate_troca_created_br_date(e.created_at)
        if bd is not None and d0 <= bd <= d1:
            out.append(e)
    return out


def _mate_troca_fmt_events_compact(evs: list[MateCouroTrocaLog]) -> str:
    if not evs:
        return "—"
    parts: list[str] = []
    for e in evs:
        k = (e.kind or "").strip()
        parts.append(
            f"{k}(+cx{e.qty_cx_in}+un{e.qty_un_in} → {e.pend_cx_after}/{e.pend_un_after})"
        )
    return " | ".join(parts)


def _mate_troca_last_balance_before_period(evs: list[MateCouroTrocaLog], d0: date) -> tuple[int, int] | None:
    """Último pend_after de evento com data BR estritamente anterior a ``d0``."""
    best: MateCouroTrocaLog | None = None
    best_bd: date | None = None
    for e in evs:
        bd = _mate_troca_created_br_date(e.created_at)
        if bd is None or bd >= d0:
            continue
        if best is None or bd > best_bd or (bd == best_bd and (e.id or 0) > (best.id or 0)):
            best = e
            best_bd = bd
    if best is None:
        return None
    return max(0, int(best.pend_cx_after or 0)), max(0, int(best.pend_un_after or 0))


def _mate_troca_base_v2_balances(
    session: Session,
    d0: date,
    d1: date,
    discovery_codes: list[str],
    period_break_map: dict[str, dict[str, int]],
) -> dict[str, dict[str, int]]:
    """Saldo operacional da Base V2.

    1. **Com log Mate troca** (qualquer histórico): uma única linha do tempo por código canônico
       (unifica aliases 010/10), ordenada por ``created_at``, ``id``. O saldo é o ``pend_*_after`` do
       **último** evento — equivale ao estado real após encadear todos os movimentos (inclui vários
       ``incorporacao_quebra`` em dias diferentes, ex. 06 e 07).

    2. **Sem log** mas presente em ``discovery_codes``: ainda não houve Carregar/sincronismo na Base;
       usa a soma de quebras (ChangeLog) no intervalo ``[d0,d1]`` — mesma agregação que monta a lista,
       refletindo o acumulado de quebra do período (ex. soma dos dias 06 e 07 dentro do De/Até).

    Assim não se usa mais um “último evento” por partição SQL em ``cod_produto`` bruto, que podia
    ignorar eventos em variantes de código do mesmo produto.

    O par CX/UN exibido é normalizado pelo fator de conversão do cadastro quando aplicável.
    """
    _ensure_mate_couro_troca_logs_table()
    fac_map = _mate_troca_product_factor_by_code(session)
    rows = list(
        session.exec(
            select(MateCouroTrocaLog).order_by(MateCouroTrocaLog.created_at, MateCouroTrocaLog.id)
        ).all()
    )
    by_canon: dict[str, list[MateCouroTrocaLog]] = defaultdict(list)
    for r in rows:
        ck = _normalize_numeric_product_code_key(str(r.cod_produto or ""))
        if not ck:
            continue
        by_canon[ck].append(r)

    replay_final: dict[str, tuple[int, int, str, MateCouroTrocaLog]] = {}
    for ck, evs in by_canon.items():
        last = evs[-1]
        replay_final[ck] = (
            max(0, int(last.pend_cx_after or 0)),
            max(0, int(last.pend_un_after or 0)),
            (last.kind or "").strip(),
            last,
        )

    all_codes = sorted(set(discovery_codes) | set(replay_final.keys()))
    balances: dict[str, dict[str, int]] = {}

    for ck in all_codes:
        evs = by_canon.get(ck, [])
        has_logs = len(evs) > 0

        saldo_anterior = _mate_troca_last_balance_before_period(evs, d0) if has_logs else None
        ev_period = _mate_troca_events_in_period_br(evs, d0, d1)
        ev_by_day: dict[date, list[MateCouroTrocaLog]] = defaultdict(list)
        for e in ev_period:
            bd = _mate_troca_created_br_date(e.created_at)
            if bd:
                ev_by_day[bd].append(e)

        zeramento_explicito = any((e.kind or "").strip() == "zerar" for e in evs)

        if has_logs:
            cx, un, last_kind, _ = replay_final[ck]
            balances[ck] = {"cx": cx, "un": un}
            origem = "replay_eventos_acumulado"
            if last_kind == "zerar":
                origem = "zeramento_explicito"
            elif last_kind in ("definir", "ajuste_pendente") and cx == 0 and un == 0:
                origem = "ultimo_evento_definir_zero"
        elif ck in period_break_map:
            rec = period_break_map[ck]
            cx = max(0, int(rec.get("cx", 0) or 0))
            un = max(0, int(rec.get("un", 0) or 0))
            balances[ck] = {"cx": cx, "un": un}
            origem = "acumulado_quebra_periodo_sem_log_mate"
            saldo_anterior = None
            zeramento_explicito = False
        else:
            continue

        # Logs temporários por código (validação operacional)
        day_keys = sorted(ev_by_day.keys())
        for dk in day_keys:
            logger.info(
                "[MATE TROCA V2][BACKEND] cod=%s dia_operacional=%s eventos=%s",
                ck,
                dk.isoformat(),
                _mate_troca_fmt_events_compact(ev_by_day[dk]),
            )
        logger.info(
            "[MATE TROCA V2][BACKEND] cod=%s saldo_anterior_antes_periodo=%s eventos_no_periodo=%s "
            "zeramento_explicito_na_historia=%s saldo_final_balances=%s/%s origem=%s",
            ck,
            saldo_anterior,
            _mate_troca_fmt_events_compact(ev_period),
            zeramento_explicito,
            balances[ck]["cx"],
            balances[ck]["un"],
            origem,
        )

        fac = fac_map.get(ck)
        cx0, un0 = balances[ck]["cx"], balances[ck]["un"]
        cxf, unf = _mate_troca_normalize_cx_un(cx0, un0, fac)
        balances[ck] = {"cx": cxf, "un": unf}

    return balances


def _sum_break_caixa_un_for_code_date_range(
    session: Session, canon_code: str, d_from: date, d_to: date
) -> tuple[int, int]:
    """Soma totais de quebra (caixa/unidade) no intervalo inclusive, chave numérica canônica."""
    c = _normalize_numeric_product_code_key(canon_code)
    if not c:
        return 0, 0
    if d_from > d_to:
        return 0, 0
    if (d_to - d_from).days > 365:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Intervalo maximo 366 dias.",
        )
    total_cx = 0
    total_un = 0
    cur = d_from
    while cur <= d_to:
        agg = _aggregate_break_events_by_code(session, cur)
        rec = agg.get(c, {})
        total_cx += int(rec.get("caixa", 0))
        total_un += int(rec.get("unidade", 0))
        cur += timedelta(days=1)
    return max(0, total_cx), max(0, total_un)


class MateTrocaReconcileFromBreaksBody(BaseModel):
    cod_produto: str = Field(min_length=1, max_length=120)
    date_from: str = Field(min_length=10, max_length=10)
    date_to: str = Field(min_length=10, max_length=10)


def _mate_troca_base_period_bounds(
    date_from: str | None,
    date_to: str | None,
) -> tuple[date, date]:
    """Limites De/Até para ``break_totals_period`` e para descoberta em ``mate-troca-base``."""
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    df_q = _parse_iso_date_arg(date_from) if date_from else None
    dt_q = _parse_iso_date_arg(date_to) if date_to else None
    if (df_q is None) ^ (dt_q is None):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Informe date_from e date_to juntos (YYYY-MM-DD) ou omita ambos.",
        )
    if df_q is None and dt_q is None:
        d0 = date(br_today.year, br_today.month, 1)
        d1 = br_today
    else:
        d0, d1 = df_q, dt_q
        if d0 > d1:
            d0, d1 = d1, d0
    return d0, d1


@router.get("/mate-troca-pending-by-product")
def get_mate_troca_pending_by_product(
    date_from: str | None = Query(
        default=None,
        description="Início do intervalo (YYYY-MM-DD) para break_totals_period, inclusive.",
    ),
    date_to: str | None = Query(
        default=None,
        description="Fim do intervalo (YYYY-MM-DD) para break_totals_period, inclusive.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Pendente CX/UN por produto a partir do último evento gravado no servidor.

    Inclui chegadas, ajustes, zeramentos e incorporacao_quebra (Carregar dia na Base de Troca).
    Usa o evento mais recente em ``created_at`` (desempate por ``id``), não só ``max(id)``,
    para refletir a ordem operacional mesmo se houver inserções fora de sequência.

    ``break_totals`` é soma histórica global (ChangeLog, CIA Mate couro) — **legado**; não usar como saldo
    operacional nem como coluna Troca na Análise de Contagem (o cliente usa ``GET /audit/mate-troca-base-v2``).

    ``break_totals_period`` é a soma das quebras no intervalo (ChangeLog) — **não** é o saldo da Base de Troca.
    A Base de Troca operacional deve usar ``GET /audit/mate-troca-base-v2`` (``balances`` + ``discovery_codes``)
    ou, no contrato antigo, ``GET /audit/mate-troca-base`` (``pending`` + ``discovery_codes``).
    """
    d0, d1 = _mate_troca_base_period_bounds(date_from, date_to)
    break_totals_period = _mate_couro_break_totals_date_range(session, d0, d1)
    return {
        "pending": _mate_troca_pending_product_map(session),
        "break_totals": _mate_couro_break_totals_all_time(session),
        "break_totals_period": break_totals_period,
    }


@router.get("/mate-troca-base")
def get_mate_troca_base(
    date_from: str | None = Query(
        default=None,
        description="Início do intervalo para descoberta de códigos com quebra (YYYY-MM-DD), inclusive.",
    ),
    date_to: str | None = Query(
        default=None,
        description="Fim do intervalo para descoberta de códigos com quebra (YYYY-MM-DD), inclusive.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Payload enxuto para a tela **Base de Troca** (operacional).

    - ``pending``: única fonte de verdade do saldo CX/UN que Chegada / Saldo / Zerar alteram.
    - ``discovery_codes``: códigos canônicos com quebra Mate couro no período (só presença no intervalo),
      para incluir na lista produtos com saldo 0 que ainda precisam de **Carregar** — **sem** expor
      totais de quebra neste endpoint.
    """
    d0, d1 = _mate_troca_base_period_bounds(date_from, date_to)
    pending = _mate_troca_pending_product_map(session)
    period_map = _mate_couro_break_totals_date_range(session, d0, d1)
    discovery_codes = sorted(period_map.keys())
    return {
        "pending": pending,
        "discovery_codes": discovery_codes,
        "period_from": d0.isoformat(),
        "period_to": d1.isoformat(),
    }


@router.get("/mate-troca-base-v2")
def get_mate_troca_base_v2(
    date_from: str | None = Query(
        default=None,
        description="Início do intervalo para descoberta de códigos com quebra (YYYY-MM-DD), inclusive.",
    ),
    date_to: str | None = Query(
        default=None,
        description="Fim do intervalo para descoberta de códigos com quebra (YYYY-MM-DD), inclusive.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Contrato explícito para a Base de Troca (UI V2).

    - ``discovery_codes``: códigos com quebra Mate couro no período (ChangeLog) — montagem da lista.
    - ``balances``: saldo operacional CX/UN. Com linha do tempo em ``mate_couro_troca_logs``, usa **replay**
      cronológico por código canônico (unifica aliases), refletindo o encadeamento de todos os eventos —
      inclusive várias ``incorporacao_quebra`` em dias distintos dentro do período. Sem log Mate troca para o
      código: usa a soma de quebras ChangeLog em ``[De, Até]`` até o Carregar sincronizar.

    A Análise de Contagem continua usando ``mate-troca-pending-by-product`` / espelho em cliente para a coluna
    Troca; esta rota é a fonte da Base V2.
    """
    d0, d1 = _mate_troca_base_period_bounds(date_from, date_to)
    period_map = _mate_couro_break_totals_date_range(session, d0, d1)
    discovery_codes = sorted(period_map.keys())
    balances = _mate_troca_base_v2_balances(session, d0, d1, discovery_codes, period_map)
    return {
        "schema_version": 2,
        "period_from": d0.isoformat(),
        "period_to": d1.isoformat(),
        "discovery_codes": discovery_codes,
        "balances": balances,
    }


@router.post("/mate-troca-reconcile-from-breaks")
def mate_troca_reconcile_pending_from_break_sum(
    body: MateTrocaReconcileFromBreaksBody,
    session: Session = Depends(get_session),
    user: User = Depends(require_roles("administrativo", "admin")),
) -> dict:
    """Define o pendente como a soma das quebras no intervalo (um evento ``definir``).

    Não substitui chegadas já registradas na operação: apenas alinha o saldo ao que consta
    nos lançamentos de quebra dos dias informados (mesma agregação da tela Quebra).
    """
    d0 = _parse_iso_date_arg(body.date_from)
    d1 = _parse_iso_date_arg(body.date_to)
    if d0 is None or d1 is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Datas invalidas (use YYYY-MM-DD).",
        )
    if d0 > d1:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="date_from nao pode ser maior que date_to.",
        )
    canon = _canonical_mate_couro_cod(session, body.cod_produto)
    tgt_cx, tgt_un = _sum_break_caixa_un_for_code_date_range(session, canon, d0, d1)
    fac_map = _mate_troca_product_factor_by_code(session)
    fac = fac_map.get(canon)
    tgt_cx, tgt_un = _mate_troca_normalize_cx_un(tgt_cx, tgt_un, fac)

    pending_map = _mate_troca_pending_product_map(session)
    cur = pending_map.get(canon) or {"cx": 0, "un": 0}
    before_cx = int(cur.get("cx", 0))
    before_un = int(cur.get("un", 0))

    base_out = {
        "ok": True,
        "cod_produto": canon,
        "pending_before": {"cx": before_cx, "un": before_un},
        "pending_target_from_breaks": {"cx": tgt_cx, "un": tgt_un},
        "date_from": d0.isoformat(),
        "date_to": d1.isoformat(),
    }

    if before_cx == tgt_cx and before_un == tgt_un:
        return {**base_out, "skipped": True, "message": "Pendente ja igual a soma das quebras no intervalo."}

    cid = f"reconcile-br-v1-{uuid.uuid4().hex}"[:100]
    ev = MateTrocaEventInput(
        client_event_id=cid,
        kind="definir",
        cod_produto=canon,
        qty_cx_in=tgt_cx,
        qty_un_in=tgt_un,
        pend_cx_before=before_cx,
        pend_un_before=before_un,
        pend_cx_after=tgt_cx,
        pend_un_after=tgt_un,
        excess_cx=0,
        excess_un=0,
        device_name="reconcile-from-breaks",
    )
    ex_cx, ex_un = _validate_mate_troca_payload(ev, fac)
    row = MateCouroTrocaLog(
        client_event_id=ev.client_event_id,
        kind="definir",
        cod_produto=canon,
        qty_cx_in=int(ev.qty_cx_in),
        qty_un_in=int(ev.qty_un_in),
        pend_cx_before=int(ev.pend_cx_before),
        pend_un_before=int(ev.pend_un_before),
        pend_cx_after=int(ev.pend_cx_after),
        pend_un_after=int(ev.pend_un_after),
        excess_cx=int(ex_cx),
        excess_un=int(ex_un),
        device_name="reconcile-from-breaks",
        actor_username=user.username,
        created_at=datetime.now(timezone.utc),
    )
    session.add(row)
    session.commit()
    return {
        **base_out,
        "skipped": False,
        "message": "Pendente atualizado (definir) pela soma das quebras no intervalo.",
        "client_event_id": cid,
    }


def _mate_troca_iso_utc(dt_val: datetime | None) -> str | None:
    if dt_val is None:
        return None
    u = dt_val if dt_val.tzinfo else dt_val.replace(tzinfo=timezone.utc)
    return u.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _mate_troca_batch_events_backward(session: Session, close: MateCouroTrocaLog) -> list[MateCouroTrocaLog]:
    if int(close.pend_cx_after or 0) != 0 or int(close.pend_un_after or 0) != 0:
        return []
    cod = str(close.cod_produto or "")
    prior = list(
        session.exec(
            select(MateCouroTrocaLog)
            .where(
                MateCouroTrocaLog.cod_produto == cod,
                MateCouroTrocaLog.created_at <= close.created_at,
            )
            .order_by(MateCouroTrocaLog.created_at.desc(), MateCouroTrocaLog.id.desc())
        ).all()
    )
    segment_rev: list[MateCouroTrocaLog] = []
    hit_close = False
    for r in prior:
        if not hit_close:
            if r.id != close.id:
                continue
            segment_rev.append(r)
            hit_close = True
            continue
        if int(r.pend_cx_after or 0) == 0 and int(r.pend_un_after or 0) == 0:
            break
        segment_rev.append(r)
    return list(reversed(segment_rev))


def _mate_logs_to_event_dicts(session: Session, rows: list[MateCouroTrocaLog]) -> list[dict]:
    codes = list({str(r.cod_produto or "") for r in rows if r.cod_produto})
    desc_map: dict[str, str] = {}
    if codes:
        prods = list(session.exec(select(Product).where(Product.cod_produto.in_(codes))).all())
        for p in prods:
            cc = _normalize_item_code(p.cod_produto or "")
            if cc:
                desc_map[cc] = (p.cod_grup_descricao or "").strip()
    mt_logins = {
        (r.actor_username or "").strip()
        for r in rows
        if r.actor_username and str(r.actor_username).strip()
    }
    mt_name_map = _display_name_map_for_logins(session, mt_logins)
    out = []
    for r in rows:
        c = str(r.cod_produto or "")
        out.append(
            {
                "id": r.id,
                "client_event_id": r.client_event_id,
                "kind": r.kind,
                "cod_produto": c,
                "product_desc": desc_map.get(c) or None,
                "qty_cx_in": int(r.qty_cx_in),
                "qty_un_in": int(r.qty_un_in),
                "pend_cx_before": int(r.pend_cx_before),
                "pend_un_before": int(r.pend_un_before),
                "pend_cx_after": int(r.pend_cx_after),
                "pend_un_after": int(r.pend_un_after),
                "excess_cx": int(r.excess_cx),
                "excess_un": int(r.excess_un),
                "device_name": r.device_name,
                "actor_username": (
                    mt_name_map.get((r.actor_username or "").strip(), r.actor_username)
                    if r.actor_username
                    else None
                ),
                "created_at": _mate_troca_iso_utc(r.created_at),
            }
        )
    return out


@router.get("/mate-troca-batches")
def list_mate_troca_batches(
    limit: int = Query(150, ge=1, le=400),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Trocas encerradas (pendente zerado no servidor), por produto Mate couro."""
    _ensure_mate_couro_troca_logs_table()
    rows_raw = list(
        session.exec(
            select(MateCouroTrocaLog)
            .order_by(MateCouroTrocaLog.created_at.desc(), MateCouroTrocaLog.id.desc())
            .limit(MAX_MATE_TROCA_BATCH_SCAN)
        ).all()
    )
    min_dt = datetime.min.replace(tzinfo=timezone.utc)

    def _row_sort_key(r: MateCouroTrocaLog):
        ca = r.created_at
        if ca is None:
            return (str(r.cod_produto or ""), min_dt, r.id or 0)
        u = ca if ca.tzinfo else ca.replace(tzinfo=timezone.utc)
        return (str(r.cod_produto or ""), u, r.id or 0)

    rows_raw.sort(key=_row_sort_key)
    closed_segments: list[list[MateCouroTrocaLog]] = []
    for _cod, git in groupby(rows_raw, key=lambda r: str(r.cod_produto or "")):
        cur: list[MateCouroTrocaLog] = []
        for ev in git:
            cur.append(ev)
            if int(ev.pend_cx_after or 0) == 0 and int(ev.pend_un_after or 0) == 0:
                closed_segments.append(cur)
                cur = []
    closed_segments.sort(
        key=lambda seg: seg[-1].created_at or min_dt,
        reverse=True,
    )
    closed_segments = closed_segments[:limit]
    summaries = []
    for seg in closed_segments:
        close = seg[-1]
        cid = close.id
        if cid is None:
            continue
        c = str(close.cod_produto or "")
        first = seg[0]
        sum_cx = sum(int(e.qty_cx_in or 0) for e in seg)
        sum_un = sum(int(e.qty_un_in or 0) for e in seg)
        summaries.append(
            {
                "batch_code": f"T-{int(cid):06d}",
                "close_log_id": int(cid),
                "cod_produto": c,
                "opened_at": _mate_troca_iso_utc(first.created_at),
                "closed_at": _mate_troca_iso_utc(close.created_at),
                "event_count": len(seg),
                "sum_qty_cx_in": sum_cx,
                "sum_qty_un_in": sum_un,
                "closing_kind": str(close.kind or "").strip(),
                "_closing_actor_login": (close.actor_username or "").strip() or None,
            }
        )
    codes = sorted({s["cod_produto"] for s in summaries if s.get("cod_produto")})
    desc_map: dict[str, str] = {}
    if codes:
        prods = list(session.exec(select(Product).where(Product.cod_produto.in_(codes))).all())
        for p in prods:
            cc = _normalize_item_code(p.cod_produto or "")
            if cc:
                desc_map[cc] = (p.cod_grup_descricao or "").strip()
    close_actor_logins = {
        s["_closing_actor_login"]
        for s in summaries
        if s.get("_closing_actor_login")
    }
    close_name_map = _display_name_map_for_logins(session, close_actor_logins)
    for s in summaries:
        s["product_desc"] = desc_map.get(s["cod_produto"]) or None
        lu = s.pop("_closing_actor_login", None)
        if lu:
            s["closed_by"] = close_name_map.get(lu, lu)
        else:
            s["closed_by"] = None
    return {"count": len(summaries), "batches": summaries}


@router.get("/mate-troca-batches/by-close/{close_log_id}")
def get_mate_troca_batch_by_close(
    close_log_id: int,
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """Detalhe de uma troca encerrada (lançamentos até zerar o pendente)."""
    _ensure_mate_couro_troca_logs_table()
    close = session.get(MateCouroTrocaLog, close_log_id)
    if close is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro nao encontrado.")
    if int(close.pend_cx_after or 0) != 0 or int(close.pend_un_after or 0) != 0:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Este evento nao encerra pendente zerado.",
        )
    segment = _mate_troca_batch_events_backward(session, close)
    if not segment:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Lote vazio.")
    events = _mate_logs_to_event_dicts(session, segment)
    lu_close = (close.actor_username or "").strip()
    cb_map = (
        _display_name_map_for_logins(session, {lu_close})
        if lu_close
        else {}
    )
    closed_by = cb_map.get(lu_close, lu_close) if lu_close else None
    return {
        "batch_code": f"T-{int(close_log_id):06d}",
        "close_log_id": int(close_log_id),
        "cod_produto": str(close.cod_produto or ""),
        "event_count": len(events),
        "closed_by": closed_by,
        "events": events,
    }


# ─────────────────────────────────────────────────────────────────
#  BI de Quebras — Dashboard
# ─────────────────────────────────────────────────────────────────


def _break_events_in_range(session: Session, d_from: date, d_to: date) -> list[dict]:
    """Lê todos os break_events no intervalo [d_from, d_to] do ChangeLog."""
    logs = list(
        session.exec(
            select(ChangeLog).where(
                ChangeLog.entity_name == "stock_break",
                ChangeLog.action == "break_event",
            )
        ).all()
    )
    out: list[dict] = []
    for log in logs:
        payload = log.payload if isinstance(log.payload, dict) else {}
        od = payload.get("operational_date")
        if od:
            try:
                d_op = date.fromisoformat(str(od)[:10])
            except ValueError:
                continue
        else:
            br_d = _brazil_date_from_observed_at(str(payload.get("observed_at") or ""))
            if br_d is None:
                continue
            d_op = br_d
        if d_op < d_from or d_op > d_to:
            continue
        item_code = str(payload.get("item_code") or "")
        code = _normalize_numeric_product_code_key(item_code)
        if not code:
            continue
        try:
            qty = int(payload.get("quantity", 0))
        except Exception:
            qty = 0
        if qty == 0:
            continue
        ct = _extract_count_type(item_code)
        out.append(
            {
                "cod_produto": code,
                "qty": qty,
                "qty_type": ct,
                "operational_date": d_op.isoformat(),
                "reason": (payload.get("reason") or "").strip() or None,
                "actor": (log.actor or "").strip() or None,
            }
        )
    return out


@router.get("/bi-quebras")
def bi_quebras_dashboard(
    date_from: str | None = Query(
        default=None,
        description="Início do intervalo YYYY-MM-DD (America/Sao_Paulo). Padrão: 30 dias atrás.",
    ),
    date_to: str | None = Query(
        default=None,
        description="Fim do intervalo YYYY-MM-DD (America/Sao_Paulo). Padrão: hoje.",
    ),
    cia_scope: str = Query(
        default=BI_QUEBRAS_CIA_SCOPE_ALL,
        description="Filtro visual por CIA: todas, mate-couro, outras.",
    ),
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
) -> dict:
    """
    Dashboard BI de Quebras: agrega eventos de quebra no intervalo e cruza com preço/categoria
    do cadastro de produto para calcular impacto financeiro estimado.

    Retorna:
    - ``summary``: totais gerais (produtos únicos, total_cx, total_un, prejuizo_brl estimado).
    - ``by_day``: série temporal de prejuízo estimado por dia operacional.
    - ``top_products``: top-N produtos por prejuízo estimado (curva ABC parcial).
    - ``by_company``: quebras por CIA (% e R$) e produtos; omite CIA sem prejuízo e sem CX/UN.
    - ``by_category``: quebras agrupadas por cod_grup_segmento (% e R$).
    - ``by_reason``: quebras agrupadas por motivo (% e R$).
    - ``products_without_price``: códigos com quebra mas sem preço cadastrado.
    """
    normalized_cia_scope = _normalize_bi_quebras_cia_scope(cia_scope)
    br_today = datetime.now(timezone.utc).astimezone(_BR).date()
    d_to_val = _parse_iso_date_arg(date_to) if date_to else br_today
    if d_to_val is None:
        d_to_val = br_today
    d_from_default = d_to_val - timedelta(days=29)
    d_from_val = _parse_iso_date_arg(date_from) if date_from else d_from_default
    if d_from_val is None:
        d_from_val = d_from_default
    if d_from_val > d_to_val:
        d_from_val, d_to_val = d_to_val, d_from_val
    if (d_to_val - d_from_val).days > 366:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Intervalo maximo 366 dias para BI de quebras.",
        )

    # 1) Coletar eventos do intervalo
    raw_events = _break_events_in_range(session, d_from_val, d_to_val)

    # 2) Buscar catálogo de produtos referenciados
    codes_set = {e["cod_produto"] for e in raw_events}
    product_map: dict[str, Any] = {}
    if codes_set:
        prods = list(session.exec(select(Product).where(Product.cod_produto.in_(list(codes_set)))).all())
        for p in prods:
            c = _normalize_numeric_product_code_key(str(p.cod_produto or ""))
            if c:
                product_map[c] = {
                    "descricao": (p.cod_grup_descricao or "").strip(),
                    "segmento": (p.cod_grup_segmento or "").strip() or "Sem segmento",
                    "marca": (p.cod_grup_marca or "").strip() or "Sem marca",
                    "cia": (p.cod_grup_cia or "").strip() or "Sem CIA",
                    "price": float(p.price) if p.price is not None else None,
                    "conversion_factor": float(p.conversion_factor) if p.conversion_factor is not None else None,
                }

    if normalized_cia_scope != BI_QUEBRAS_CIA_SCOPE_ALL:
        raw_events = [
            e
            for e in raw_events
            if _matches_bi_quebras_cia_scope(product_map.get(e["cod_produto"], {}).get("cia"), normalized_cia_scope)
        ]

    def _estimate_loss(cod: str, qty_cx: int, qty_un: int) -> float | None:
        """Prejuízo estimado em R$ (unidades equivalentes × preço unitário)."""
        info = product_map.get(cod)
        if not info:
            return None
        price = info.get("price")
        if price is None or price <= 0:
            return None
        conv = info.get("conversion_factor") or 1.0
        total_un_equiv = qty_un + qty_cx * conv
        return round(total_un_equiv * price, 2)

    # 3) Agregar quebras por produto
    by_code: dict[str, dict] = {}
    for e in raw_events:
        cod = e["cod_produto"]
        if cod not in by_code:
            by_code[cod] = {"cx": 0, "un": 0}
        if e["qty_type"] == "caixa":
            by_code[cod]["cx"] += e["qty"]
        else:
            by_code[cod]["un"] += e["qty"]

    products_without_price: list[str] = []
    for cod, rec in by_code.items():
        info = product_map.get(cod, {})
        loss = _estimate_loss(cod, rec["cx"], rec["un"])
        rec["loss_brl"] = loss
        rec["descricao"] = info.get("descricao", cod)
        rec["segmento"] = info.get("segmento", "Sem segmento")
        rec["marca"] = info.get("marca", "Sem marca")
        rec["cia"] = _normalize_bi_quebras_cia_label(info.get("cia"))
        if info.get("price") is None:
            products_without_price.append(cod)

    # 4) Série temporal por dia operacional
    by_day_raw: dict[str, dict] = {}
    for e in raw_events:
        d_key = e["operational_date"]
        if d_key not in by_day_raw:
            by_day_raw[d_key] = {}
        cod = e["cod_produto"]
        if cod not in by_day_raw[d_key]:
            by_day_raw[d_key][cod] = {"cx": 0, "un": 0}
        if e["qty_type"] == "caixa":
            by_day_raw[d_key][cod]["cx"] += e["qty"]
        else:
            by_day_raw[d_key][cod]["un"] += e["qty"]

    by_day: list[dict] = []
    cur = d_from_val
    while cur <= d_to_val:
        dk = cur.isoformat()
        day_prods = by_day_raw.get(dk, {})
        day_loss: float = 0.0
        day_items = len(day_prods)
        has_any_price = False
        for cod2, rec2 in day_prods.items():
            l = _estimate_loss(cod2, rec2["cx"], rec2["un"])
            if l is not None:
                day_loss += l
                has_any_price = True
        by_day.append({
            "date": dk,
            "items_with_break": day_items,
            "loss_brl": round(day_loss, 2) if has_any_price else None,
        })
        cur += timedelta(days=1)

    # 5) Top produtos por prejuízo estimado
    ranked = sorted(
        by_code.items(),
        key=lambda kv: (kv[1].get("loss_brl") or 0),
        reverse=True,
    )
    top_products = [
        {
            "cod_produto": cod,
            "descricao": rec["descricao"],
            "cia": rec["cia"],
            "segmento": rec["segmento"],
            "cx": int(rec["cx"]),
            "un": int(rec["un"]),
            "loss_brl": rec.get("loss_brl"),
        }
        for cod, rec in ranked[:20]
    ]

    # 6) Agrupamento por segmento
    seg_agg: dict[str, dict] = {}
    for cod, rec in by_code.items():
        seg = rec["segmento"]
        if seg not in seg_agg:
            seg_agg[seg] = {"loss_brl": 0.0, "items": 0, "has_price": False}
        seg_agg[seg]["items"] += 1
        l = rec.get("loss_brl")
        if l is not None:
            seg_agg[seg]["loss_brl"] += l
            seg_agg[seg]["has_price"] = True
    total_seg_loss = sum(v["loss_brl"] for v in seg_agg.values() if v["has_price"])
    by_category = [
        {
            "segmento": seg,
            "items": v["items"],
            "loss_brl": round(v["loss_brl"], 2) if v["has_price"] else None,
            "pct": round(v["loss_brl"] / total_seg_loss * 100, 1) if total_seg_loss > 0 and v["has_price"] else None,
        }
        for seg, v in sorted(seg_agg.items(), key=lambda kv: kv[1]["loss_brl"], reverse=True)
    ]

    # 6.1) Agrupamento por CIA com lista de produtos clicável no front
    company_agg: dict[str, dict] = {}
    for cod, rec in by_code.items():
        cia = _normalize_bi_quebras_cia_label(rec.get("cia"))
        if cia not in company_agg:
            company_agg[cia] = {
                "loss_brl": 0.0,
                "items": 0,
                "total_cx": 0,
                "total_un": 0,
                "has_price": False,
                "products": [],
            }
        company_agg[cia]["items"] += 1
        company_agg[cia]["total_cx"] += int(rec["cx"])
        company_agg[cia]["total_un"] += int(rec["un"])
        loss = rec.get("loss_brl")
        if loss is not None:
            company_agg[cia]["loss_brl"] += loss
            company_agg[cia]["has_price"] = True
        company_agg[cia]["products"].append(
            {
                "cod_produto": cod,
                "descricao": rec["descricao"],
                "segmento": rec["segmento"],
                "cia": cia,
                "cx": int(rec["cx"]),
                "un": int(rec["un"]),
                "loss_brl": loss,
            }
        )

    total_company_loss = sum(v["loss_brl"] for v in company_agg.values() if v["has_price"])
    by_company = []
    for cia, v in sorted(company_agg.items(), key=lambda kv: kv[1]["loss_brl"], reverse=True):
        loss_amt = float(v["loss_brl"])
        total_cx = int(v["total_cx"])
        total_un = int(v["total_un"])
        # Sem quebra efetiva: não listar CIA (evita linhas 0 R$ · 0 CX · 0 UN).
        if loss_amt <= 0 and total_cx == 0 and total_un == 0:
            continue
        products = sorted(
            v["products"],
            key=lambda p: (
                -(p.get("loss_brl") or 0),
                -(p.get("cx") or 0),
                -(p.get("un") or 0),
                str(p.get("descricao") or "").lower(),
            ),
        )
        by_company.append(
            {
                "cia": cia,
                "items": v["items"],
                "total_cx": total_cx,
                "total_un": total_un,
                "loss_brl": round(v["loss_brl"], 2) if v["has_price"] else None,
                "pct": round(v["loss_brl"] / total_company_loss * 100, 1)
                if total_company_loss > 0 and v["has_price"]
                else None,
                "products": products,
            }
        )

    # 7) Agrupamento por motivo
    reason_agg: dict[str, dict] = {}
    for e in raw_events:
        reason = e.get("reason") or "Não informado"
        cod = e["cod_produto"]
        if reason not in reason_agg:
            reason_agg[reason] = {"count": 0, "codes": set(), "loss_brl": 0.0, "has_price": False}
        reason_agg[reason]["count"] += 1
        reason_agg[reason]["codes"].add(cod)

    # Distribuir o prejuízo proporcionalmente pelos eventos (eventos de cada motivo)
    for e in raw_events:
        cod = e["cod_produto"]
        reason = e.get("reason") or "Não informado"
        info = product_map.get(cod, {})
        price = info.get("price")
        if price is None or price <= 0:
            continue
        conv = info.get("conversion_factor") or 1.0
        n_events_this_cod = sum(1 for ev in raw_events if ev["cod_produto"] == cod)
        qty = e["qty"]
        ct = e["qty_type"]
        un_equiv = qty if ct == "unidade" else qty * conv
        loss_share = round(un_equiv * price, 4)
        if reason in reason_agg:
            reason_agg[reason]["loss_brl"] += loss_share
            reason_agg[reason]["has_price"] = True

    total_reason_loss = sum(v["loss_brl"] for v in reason_agg.values() if v["has_price"])
    by_reason = [
        {
            "reason": reason,
            "occurrences": v["count"],
            "unique_products": len(v["codes"]),
            "loss_brl": round(v["loss_brl"], 2) if v["has_price"] else None,
            "pct": round(v["loss_brl"] / total_reason_loss * 100, 1) if total_reason_loss > 0 and v["has_price"] else None,
        }
        for reason, v in sorted(reason_agg.items(), key=lambda kv: kv[1]["loss_brl"], reverse=True)
    ]

    # 8) Summary geral
    total_loss_brl = sum(
        rec["loss_brl"] for rec in by_code.values() if rec.get("loss_brl") is not None
    )
    total_cx = sum(rec["cx"] for rec in by_code.values())
    total_un = sum(rec["un"] for rec in by_code.values())

    return {
        "date_from": d_from_val.isoformat(),
        "date_to": d_to_val.isoformat(),
        "cia_scope": normalized_cia_scope,
        "summary": {
            "unique_products": len(by_code),
            "total_cx": int(total_cx),
            "total_un": int(total_un),
            "total_loss_brl": round(total_loss_brl, 2),
            "products_with_price": len(by_code) - len(products_without_price),
            "products_without_price": len(products_without_price),
        },
        "by_day": by_day,
        "top_products": top_products,
        "by_company": by_company,
        "by_category": by_category,
        "by_reason": by_reason,
        "products_without_price": products_without_price[:50],
    }
