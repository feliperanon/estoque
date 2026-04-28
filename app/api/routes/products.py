from io import BytesIO
import logging
import re
import unicodedata
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import String, and_, cast, func, or_, text
from sqlalchemy.exc import SQLAlchemyError
from sqlmodel import Session, select

from app.api.deps import require_cadastro_access, require_roles
from app.db.session import get_session
from app.models import Product, ProductHistory, User
from app.schemas.products import ProductCreate, ProductHistoryRead, ProductImportPayload, ProductRead, ProductUpdate
from app.services.audit import log_change
from app.services.bootstrap import ensure_database_ready
from app.services.imports import apply_common_source_fields
from app.services.sqlite_product_constraints import apply_sqlite_product_unique_constraints

router = APIRouter(prefix="/products", tags=["products"])
logger = logging.getLogger(__name__)


def _safe_log_change(
    session: Session,
    entity_name: str,
    entity_id: int,
    action: str,
    actor: str | None,
    payload: dict | None = None,
) -> None:
    try:
        # Isola falhas de auditoria sem contaminar a transacao principal.
        with session.begin_nested():
            log_change(session, entity_name, entity_id, action, actor, payload)
            # Forca o INSERT dentro do savepoint para capturar UndefinedTable aqui.
            session.flush()
    except Exception:
        # Auditoria nao pode derrubar operacao principal.
        logger.exception("Falha ao registrar auditoria", extra={"entity": entity_name, "action": action})


def _sanitize_product_for_response(product: Product) -> Product:
    # Compatibilidade com dados legados incompletos para evitar erro 500 de serializacao.
    sku = (product.cod_grup_sku or "").strip()
    cod_produto = (product.cod_produto or "").strip()
    descricao = (product.cod_grup_descricao or "").strip()

    if not sku:
        sku = cod_produto or str(product.id or "")
        product.cod_grup_sku = sku
    if not cod_produto:
        product.cod_produto = sku or str(product.id or "")
    if not descricao:
        product.cod_grup_descricao = sku or product.cod_produto or f"Produto {product.id or ''}".strip()

    if getattr(product, "updated_at", None) is None:
        product.updated_at = datetime.now(timezone.utc)

    for field_name in ("updated_at", "created_at", "imported_at"):
        value = getattr(product, field_name, None)
        if value is not None and getattr(value, "tzinfo", None) is None:
            setattr(product, field_name, value.replace(tzinfo=timezone.utc))
    return product


def _to_product_read(product: Product) -> ProductRead:
    safe = _sanitize_product_for_response(product)
    return ProductRead(
        id=safe.id or 0,
        cod_grup_sp=safe.cod_grup_sp,
        cod_grup_cia=safe.cod_grup_cia,
        cod_grup_tipo=safe.cod_grup_tipo,
        cod_grup_familia=safe.cod_grup_familia,
        cod_grup_segmento=safe.cod_grup_segmento,
        cod_grup_marca=safe.cod_grup_marca,
        cod_produto=(safe.cod_produto or "").strip() or (safe.cod_grup_sku or str(safe.id or "")),
        cod_grup_descricao=(safe.cod_grup_descricao or "").strip() or (safe.cod_grup_sku or "Sem descricao"),
        cod_grup_sku=(safe.cod_grup_sku or "").strip() or (safe.cod_produto or str(safe.id or "")),
        status=safe.status,
        grup_prioridade=safe.grup_prioridade,
        price=safe.price,
        conversion_factor=safe.conversion_factor,
        pallet_conversion_factor=safe.pallet_conversion_factor,
        legacy_id=safe.legacy_id,
        source_system=safe.source_system,
        imported_at=safe.imported_at,
        updated_at=safe.updated_at,
        created_at=safe.created_at,
    )


def _serialize_products(products: list[Product]) -> list[ProductRead]:
    result: list[ProductRead] = []
    for product in products:
        try:
            result.append(_to_product_read(product))
        except Exception:
            logger.exception("Registro de produto invalido ignorado", extra={"product_id": getattr(product, "id", None)})
    return result


def _drop_legacy_sku_unique_constraint(session: Session) -> None:
    """Remove unique legada em cod_grup_sku (Postgres) ou SQLite; alinha com unicidade por cod_produto."""
    bind = session.get_bind()
    if bind.dialect.name == "sqlite":
        apply_sqlite_product_unique_constraints(session)
        session.flush()
        return

    if bind.dialect.name != "postgresql":
        return

    # Remove nome conhecido (idempotente).
    session.execute(text("ALTER TABLE app_core.products DROP CONSTRAINT IF EXISTS uq_product_sku"))

    # Remove qualquer unique em cima apenas de cod_grup_sku, mesmo com outro nome.
    rows = session.execute(
        text(
            """
            SELECT c.conname
            FROM pg_constraint c
            JOIN pg_class t ON t.oid = c.conrelid
            JOIN pg_namespace n ON n.oid = t.relnamespace
            JOIN pg_attribute a ON a.attrelid = t.oid AND a.attnum = ANY(c.conkey)
            WHERE c.contype = 'u'
              AND n.nspname = 'app_core'
              AND t.relname = 'products'
            GROUP BY c.oid, c.conname
            HAVING bool_and(a.attname = 'cod_grup_sku')
            """
        )
    ).fetchall()

    for (constraint_name,) in rows:
        if not constraint_name:
            continue
        logger.warning("Removendo unique legada em SKU", extra={"constraint": constraint_name})
        session.execute(
            text(f'ALTER TABLE app_core.products DROP CONSTRAINT IF EXISTS "{constraint_name}"')
        )

    session.flush()


def _norm_header(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", (value or "").strip().lower())
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")
    return normalized


# Modelo oficial de importação (planilha BI — 8 colunas):
# Cia → cod_grup_cia | Tipo → cod_grup_tipo | Segmento → cod_grup_segmento | Marca → cod_grup_marca
# Produto → cod_grup_descricao | Codigo → cod_produto | SKU → cod_grup_sku | Custo → price | Fator → conversion_factor
HEADER_ALIASES = {
    "cod_grup_sp": "cod_grup_sp",
    "grup_sp": "cod_grup_sp",
    "cod_sp": "cod_grup_sp",
    "cod_grup_cia": "cod_grup_cia",
    "grup_cia": "cod_grup_cia",
    "cod_cia": "cod_grup_cia",
    "cod_grup_tipo": "cod_grup_tipo",
    "grup_tipo": "cod_grup_tipo",
    "cod_tipo": "cod_grup_tipo",
    "cod_grup_familia": "cod_grup_familia",
    "grup_familia": "cod_grup_familia",
    "cod_familia": "cod_grup_familia",
    "cod_grup_segmento": "cod_grup_segmento",
    "grup_segmento": "cod_grup_segmento",
    "cod_segmento": "cod_grup_segmento",
    "cod_grup_marca": "cod_grup_marca",
    "grup_marca": "cod_grup_marca",
    "cod_marca": "cod_grup_marca",
    "cod_produto": "cod_produto",
    "codigo_produto": "cod_produto",
    "codigo_do_produto": "cod_produto",
    "codigo": "cod_produto",
    "cod": "cod_produto",
    "cod_item": "cod_produto",
    "codigo_item": "cod_produto",
    "cod_interno": "cod_produto",
    "codigo_interno": "cod_produto",
    "id_produto": "cod_produto",
    "cod_grup_descricao": "cod_grup_descricao",
    "grup_descricao": "cod_grup_descricao",
    "descricao": "cod_grup_descricao",
    "descricao_do_produto": "cod_grup_descricao",
    "descricao_produto": "cod_grup_descricao",
    "nome_produto": "cod_grup_descricao",
    "nome_do_produto": "cod_grup_descricao",
    "produto": "cod_grup_descricao",
    "nome": "cod_grup_descricao",
    "item": "cod_grup_descricao",
    "denominacao": "cod_grup_descricao",
    "mercadoria": "cod_grup_descricao",
    "produto_descricao": "cod_grup_descricao",
    "cod_grup_sku": "cod_grup_sku",
    "grup_sku": "cod_grup_sku",
    "sku": "cod_grup_sku",
    "cod_sku": "cod_grup_sku",
    "codigo_sku": "cod_grup_sku",
    "referencia": "cod_grup_sku",
    "ref": "cod_grup_sku",
    "codigo_referencia": "cod_grup_sku",
    "codigo_barras": "cod_grup_sku",
    "ean": "cod_grup_sku",
    "gtin": "cod_grup_sku",
    "status": "status",
    "grup_prioridade": "grup_prioridade",
    "prioridade": "grup_prioridade",
    # Cabeçalhos curtos (planilhas operacionais)
    "cia": "cod_grup_cia",
    "tipo": "cod_grup_tipo",
    "segmento": "cod_grup_segmento",
    "marca": "cod_grup_marca",
    # Custo/preço no cadastro (campo price no modelo)
    "price": "price",
    "preco": "price",
    "preco_rs": "price",
    "preco_unitario": "price",
    "custo": "price",
    "custo_rs": "price",
    "custo_unitario": "price",
    "valor": "price",
    "valor_rs": "price",
    "valor_unitario": "price",
    # Fator CX→UN (unidades por 1 caixa)
    "conversion_factor": "conversion_factor",
    "fator_conversao": "conversion_factor",
    "fator_de_conversao": "conversion_factor",
    "unidades_por_caixa": "conversion_factor",
    "un_por_cx": "conversion_factor",
    "unidades_por_cx": "conversion_factor",
    "cx_para_un": "conversion_factor",
    # Fator PL→CX (caixas por 1 palete)
    "pallet_conversion_factor": "pallet_conversion_factor",
    "fator_conversao_palete": "pallet_conversion_factor",
    "fator_palete": "pallet_conversion_factor",
    "palete_por_caixa": "pallet_conversion_factor",
    "caixas_por_palete": "pallet_conversion_factor",
    "cx_por_pl": "pallet_conversion_factor",
    "pl_para_cx": "pallet_conversion_factor",
}

# Importação por planilha: obrigatório código do produto + nome/descrição. SKU é opcional (espelha o código se vazio).
REQUIRED_FIELDS = {"cod_produto", "cod_grup_descricao"}

_ATIVO_STATUS_SYNONYMS = frozenset(
    {
        "ativo",
        "ativado",
        "active",
        "s",
        "sim",
        "si",
        "yes",
        "y",
        "1",
        "true",
        "verdadeiro",
        "ok",
        "x",
    }
)
_INATIVO_STATUS_SYNONYMS = frozenset(
    {
        "inativo",
        "inactive",
        "n",
        "nao",
        "no",
        "0",
        "false",
        "falso",
    }
)


def _strip_accents_lower(s: str) -> str:
    n = unicodedata.normalize("NFKD", s)
    return "".join(c for c in n if not unicodedata.combining(c)).lower().strip()


def _normalize_import_status(raw: str | bool | None) -> str | None:
    """Alinha status de planilhas BI (S/N, 1/0, boolean) ao modelo ativo/inativo."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return "ativo" if raw else "inativo"
    s = str(raw).strip()
    if not s:
        return None
    key = _strip_accents_lower(s)
    if key in _ATIVO_STATUS_SYNONYMS:
        return "ativo"
    if key in _INATIVO_STATUS_SYNONYMS:
        return "inativo"
    return s


def _catalog_status_is_ativo_clause():
    """Filtro SQL: considera ativo os mesmos sinonimos aceitos na importacao."""
    col = cast(Product.status, String)
    sl = func.lower(func.trim(col))
    return or_(
        Product.status.is_(None),
        func.trim(col) == "",
        sl == "ativo",
        sl.in_(_ATIVO_STATUS_SYNONYMS),
        Product.status == "Ativo",
        Product.status == "ATIVO",
    )


def _status_pre_cadastro_clause():
    """Status de pré-cadastro (rótulos comuns e variações com/sem acento)."""
    col = cast(Product.status, String)
    sl = func.lower(func.trim(col))
    raw = Product.status
    return or_(
        sl.in_(("pre-cadastro", "pré-cadastro", "pre cadastro", "pré cadastro", "precadastro")),
        sl.like("%pre%cadastro%"),
        sl.like("%pré%cadastro%"),
        raw.ilike("%pre%cadastro%"),
        raw.ilike("%pré%cadastro%"),
    )


def _status_inativo_clause():
    """Inativo explícito (sinônimos da importação), excluindo ativos e pré-cadastro."""
    col = cast(Product.status, String)
    sl = func.lower(func.trim(col))
    inativo_ish = or_(
        sl == "inativo",
        sl.in_(tuple(_INATIVO_STATUS_SYNONYMS)),
        Product.status == "Inativo",
        Product.status == "INATIVO",
    )
    return and_(
        inativo_ish,
        func.not_(_catalog_status_is_ativo_clause()),
        func.not_(_status_pre_cadastro_clause()),
    )


def _normalize_list_status_tokens(raw: list[str] | None) -> set[str]:
    """Normaliza query params status=... para ativo | inativo | pre-cadastro."""
    if not raw:
        return set()
    out: set[str] = set()
    for item in raw:
        t = _strip_accents_lower((item or "").strip())
        if not t:
            continue
        if t in ("ativo", "ativos", "a"):
            out.add("ativo")
        elif t in ("inativo", "inativos", "i"):
            out.add("inativo")
        elif t in ("pre-cadastro", "pre cadastro", "precadastro", "p", "pre") or (
            "pre" in t and "cadastro" in t
        ):
            out.add("pre-cadastro")
    return out


def _parse_br_price(value: object) -> float | None:
    """Interpreta custo/preço: número Excel, ou texto tipo R$ 1.234,56 / 60,03."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"^(R\$|USD|\$)\s*", "", s, flags=re.IGNORECASE).strip()
    s = s.replace(" ", "")
    if not s:
        return None
    if "," in s and "." in s:
        # 1.234,56 (BR) vs 1,234.56 (US)
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2 and parts[1].isdigit():
            int_part = parts[0].replace(".", "")
            s = f"{int_part}.{parts[1]}"
        else:
            s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_conversion_factor_in_row(row_data: dict[str, str | float | None]) -> None:
    """Converte coluna de fator de conversão (UN por CX) para float positivo."""
    raw = row_data.get("conversion_factor")
    if raw is None:
        return
    if isinstance(raw, str) and not raw.strip():
        row_data.pop("conversion_factor", None)
        return
    if isinstance(raw, bool):
        row_data.pop("conversion_factor", None)
        return
    if isinstance(raw, (int, float)):
        v = float(raw)
    else:
        s = str(raw).strip().replace(",", ".")
        try:
            v = float(s)
        except ValueError:
            row_data.pop("conversion_factor", None)
            return
    if v <= 0 or v != v:  # exclui NaN
        row_data.pop("conversion_factor", None)
        return
    row_data["conversion_factor"] = v


def _coerce_pallet_conversion_factor_in_row(row_data: dict[str, str | float | None]) -> None:
    """Converte coluna de fator palete→caixa (CX por PL) para float positivo."""
    raw = row_data.get("pallet_conversion_factor")
    if raw is None:
        return
    if isinstance(raw, str) and not raw.strip():
        row_data.pop("pallet_conversion_factor", None)
        return
    if isinstance(raw, bool):
        row_data.pop("pallet_conversion_factor", None)
        return
    if isinstance(raw, (int, float)):
        v = float(raw)
    else:
        s = str(raw).strip().replace(",", ".")
        try:
            v = float(s)
        except ValueError:
            row_data.pop("pallet_conversion_factor", None)
            return
    if v <= 0 or v != v:
        row_data.pop("pallet_conversion_factor", None)
        return
    row_data["pallet_conversion_factor"] = v


def _coerce_price_in_row(row_data: dict[str, str | float | None]) -> None:
    """Converte coluna de custo/preço da planilha para float."""
    raw = row_data.get("price")
    if raw is None or (isinstance(raw, str) and not str(raw).strip()):
        row_data.pop("price", None)
        return
    parsed = _parse_br_price(raw)
    if parsed is None:
        row_data.pop("price", None)
    else:
        row_data["price"] = parsed


def _normalize_codigo_import(row_data: dict[str, str | None]) -> None:
    """Evita cod_produto '140.0' quando o Excel grava código como float."""
    raw = row_data.get("cod_produto")
    if raw is None:
        return
    s = str(raw).strip()
    if not s:
        return
    if re.fullmatch(r"-?\d+\.0+", s):
        row_data["cod_produto"] = s.split(".")[0]
        return
    try:
        num = float(s.replace(",", "."))
        if num == int(num) and abs(num) < 1e15:
            row_data["cod_produto"] = str(int(num))
            return
    except ValueError:
        pass
    row_data["cod_produto"] = s


def _fill_import_row_defaults(row_data: dict[str, str | None]) -> None:
    """Preenche codigo/SKU/descricao quando a planilha traz só codigo+nome ou colunas BI parciais."""
    cod = (row_data.get("cod_produto") or "").strip()
    sku = (row_data.get("cod_grup_sku") or "").strip()
    desc = (row_data.get("cod_grup_descricao") or "").strip()
    if not sku and cod:
        row_data["cod_grup_sku"] = cod
        sku = cod
    if not cod and sku:
        row_data["cod_produto"] = sku
        cod = sku
    if not desc:
        row_data["cod_grup_descricao"] = (cod or sku or "").strip() or None


def _map_headers(raw_headers: tuple) -> tuple[list[str | None], int]:
    mapped: list[str | None] = []
    for header in raw_headers:
        normalized = _norm_header(str(header or ""))
        mapped.append(HEADER_ALIASES.get(normalized))

    # score por quantidade de campos reconhecidos (sem duplicidade)
    recognized = {item for item in mapped if item}
    return mapped, len(recognized)


def _parse_optional_price_query(q: str) -> float | None:
    """Interpreta texto como valor monetário para busca pelo campo custo (price)."""
    raw = (q or "").strip()
    if not raw:
        return None
    raw = re.sub(r"^\s*R\$\s*", "", raw, flags=re.I).strip()
    if not raw:
        return None
    if re.search(r",\d{1,2}$", raw):
        normalized = raw.replace(".", "").replace(",", ".")
    else:
        normalized = raw.replace(",", ".")
    try:
        val = float(normalized)
    except ValueError:
        return None
    if val != val or val < 0 or val > 1e12:  # exclui NaN
        return None
    return round(val, 2)


def _product_list_search_clause(q: str):
    """Busca em código, descrição, SKU, Cia, Tipo, Segmento, Marca e custo (valor numérico)."""
    qn = (q or "").strip()
    if not qn:
        return None
    ql = qn.lower()
    parts = [
        func.lower(func.coalesce(Product.cod_produto, "")).contains(ql),
        func.lower(Product.cod_grup_descricao).contains(ql),
        func.lower(Product.cod_grup_sku).contains(ql),
        func.lower(func.coalesce(Product.cod_grup_cia, "")).contains(ql),
        func.lower(func.coalesce(Product.cod_grup_tipo, "")).contains(ql),
        func.lower(func.coalesce(Product.cod_grup_segmento, "")).contains(ql),
        func.lower(func.coalesce(Product.cod_grup_marca, "")).contains(ql),
    ]
    pv = _parse_optional_price_query(qn)
    if pv is not None:
        parts.append(and_(Product.price.isnot(None), func.abs(Product.price - pv) < 0.005))
    return or_(*parts)


@router.get("", response_model=list[ProductRead])
def list_products(
    session: Session = Depends(get_session),
    _: User = Depends(require_cadastro_access),
    q: str | None = Query(default=None),
    limit: int = Query(default=500, ge=1, le=20000),
    status: list[str] | None = Query(default=None),
    cia: str | None = Query(default=None),
    segmento: str | None = Query(default=None),
    marca: str | None = Query(default=None),
) -> list[ProductRead]:
    statement = select(Product)
    q_stripped = (q or "").strip()
    if q_stripped:
        clause = _product_list_search_clause(q_stripped)
        if clause is not None:
            statement = statement.where(clause)

    if (cia or "").strip():
        statement = statement.where(Product.cod_grup_cia == cia.strip())
    if (segmento or "").strip():
        statement = statement.where(Product.cod_grup_segmento == segmento.strip())
    if (marca or "").strip():
        statement = statement.where(Product.cod_grup_marca == marca.strip())

    tokens = _normalize_list_status_tokens(status)
    if tokens:
        if tokens == {"ativo", "inativo", "pre-cadastro"}:
            pass
        else:
            parts = []
            if "ativo" in tokens:
                parts.append(_catalog_status_is_ativo_clause())
            if "inativo" in tokens:
                parts.append(_status_inativo_clause())
            if "pre-cadastro" in tokens:
                parts.append(_status_pre_cadastro_clause())
            if parts:
                statement = statement.where(or_(*parts))

    try:
        products = list(session.exec(statement.order_by(Product.cod_grup_descricao).limit(limit)).all())
    except SQLAlchemyError:
        session.rollback()
        ensure_database_ready()
        products = list(session.exec(statement.order_by(Product.cod_grup_descricao).limit(limit)).all())
    except Exception:
        session.rollback()
        logger.exception("Falha inesperada ao listar produtos")
        return []

    return _serialize_products(products)


@router.get("/catalog", response_model=list[ProductRead])
def list_products_catalog(
    session: Session = Depends(get_session),
    _: User = Depends(require_roles("conferente", "administrativo", "admin")),
    q: str | None = Query(default=None),
    status_filter: str = Query(default="todos", alias="status"),
    cia: str | None = Query(default=None, description="Filtra por cod_grup_cia exato."),
    limit: int = Query(default=500, ge=1, le=20000),
) -> list[ProductRead]:
    statement = select(Product)
    normalized_status = (status_filter or "todos").strip().lower()
    if normalized_status == "ativo":
        # Legado + planilhas BI (S, 1, SIM, boolean como texto).
        statement = statement.where(_catalog_status_is_ativo_clause())
    elif normalized_status != "todos":
        statement = statement.where(
            or_(
                Product.status == normalized_status,
                Product.status == normalized_status.capitalize(),
                Product.status == normalized_status.upper()
            )
        )
    if (cia or "").strip():
        statement = statement.where(Product.cod_grup_cia == cia.strip())
    if q:
        statement = statement.where(
            Product.cod_produto.contains(q)
            | Product.cod_grup_descricao.contains(q)
            | Product.cod_grup_marca.contains(q)
            | Product.cod_grup_sku.contains(q),
        )

    try:
        products = list(session.exec(statement.order_by(Product.cod_produto).limit(limit)).all())
    except SQLAlchemyError:
        session.rollback()
        ensure_database_ready()
        products = list(session.exec(statement.order_by(Product.cod_produto).limit(limit)).all())
    except Exception:
        session.rollback()
        logger.exception("Falha inesperada ao listar catalogo de produtos")
        return []
    return _serialize_products(products)


@router.post("", response_model=ProductRead)
def create_product(
    payload: ProductCreate,
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> Product:
    _drop_legacy_sku_unique_constraint(session)
    cod = (payload.cod_produto or "").strip()
    if not cod:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Codigo do produto obrigatorio")
    existing = session.exec(select(Product).where(Product.cod_produto == cod)).first()
    if existing:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Codigo do produto ja cadastrado")

    product = Product.model_validate(payload)
    apply_common_source_fields(product, payload.legacy_id, payload.source_system or "manual")
    session.add(product)
    session.flush()
    _safe_log_change(session, "products", product.id or 0, "create", user.username, payload.model_dump())
    session.commit()
    session.refresh(product)
    return product


@router.post("/import", response_model=dict)
def import_products_payload(
    payload: ProductImportPayload,
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> dict:
    _drop_legacy_sku_unique_constraint(session)
    created = 0
    updated = 0

    for row in payload.rows:
        cod = (row.cod_produto or "").strip()
        if not cod:
            continue
        existing = session.exec(select(Product).where(Product.cod_produto == cod)).first()
        if existing:
            data = row.model_dump(exclude={"legacy_id", "source_system"})
            for key, value in data.items():
                setattr(existing, key, value)
            apply_common_source_fields(existing, row.legacy_id, row.source_system or "excel")
            updated += 1
            continue

        product = Product.model_validate(row)
        apply_common_source_fields(product, row.legacy_id, row.source_system or "excel")
        session.add(product)
        created += 1

    session.flush()
    _safe_log_change(
        session,
        "products",
        0,
        "import",
        user.username,
        {"created": created, "updated": updated, "rows": len(payload.rows)},
    )
    session.commit()
    return {"created": created, "updated": updated, "rows": len(payload.rows)}


@router.post("/import-excel", response_model=dict)
async def import_products_excel(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> dict:
    _drop_legacy_sku_unique_constraint(session)
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Envie um arquivo .xlsx")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Arquivo vazio")

    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Planilha sem cabecalho")

        header_row_index = -1
        mapped_headers: list[str | None] = []
        best_score = 0

        # Algumas planilhas trazem titulo/descricao antes do cabecalho.
        for idx, row in enumerate(all_rows[:30]):
            current_mapped, score = _map_headers(row)
            if score > best_score:
                best_score = score
                mapped_headers = current_mapped
                header_row_index = idx

        if best_score < 2 or header_row_index < 0:
            first_line_preview = [str(v or "").strip() for v in all_rows[0][:12]]
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    "Nao foi possivel reconhecer colunas de produto no cabecalho. "
                    f"Cabecalho lido: {first_line_preview}"
                ),
            )

        created = 0
        updated = 0
        ignored = 0
        failed = 0
        created_in_batch: dict[str, Product] = {}
        product_codes_touched: set[str] = set()

        for row in all_rows[header_row_index + 1 :]:
            row_data: dict[str, str | None] = {}
            for idx, value in enumerate(row):
                mapped = mapped_headers[idx] if idx < len(mapped_headers) else None
                if not mapped:
                    continue
                if value is None:
                    row_data[mapped] = None
                else:
                    row_data[mapped] = str(value).strip()

            if not any(row_data.values()):
                continue

            _normalize_codigo_import(row_data)
            _coerce_price_in_row(row_data)
            _coerce_conversion_factor_in_row(row_data)
            _coerce_pallet_conversion_factor_in_row(row_data)
            _fill_import_row_defaults(row_data)
            if "status" in row_data and row_data.get("status") is not None:
                row_data["status"] = _normalize_import_status(row_data["status"])

            cod_produto = (row_data.get("cod_produto") or "").strip()
            descricao = (row_data.get("cod_grup_descricao") or "").strip()
            if not cod_produto or not descricao:
                ignored += 1
                continue

            row_data["cod_produto"] = cod_produto
            row_data["cod_grup_descricao"] = descricao
            # SKU não é chave de negócio: se a planilha não tiver coluna SKU, usa o mesmo código do produto.
            sku = (row_data.get("cod_grup_sku") or "").strip() or cod_produto
            row_data["cod_grup_sku"] = sku

            try:
                # Mesmo codigo de produto no arquivo: uma unica linha de cadastro (atualiza a staged).
                if cod_produto in created_in_batch:
                    staged = created_in_batch[cod_produto]
                    for key, value in row_data.items():
                        setattr(staged, key, value)
                    apply_common_source_fields(staged, None, "excel")
                    updated += 1
                    product_codes_touched.add(cod_produto)
                    continue

                existing = session.exec(select(Product).where(Product.cod_produto == cod_produto)).first()

                if existing:
                    for key, value in row_data.items():
                        setattr(existing, key, value)
                    apply_common_source_fields(existing, None, "excel")
                    updated += 1
                    product_codes_touched.add(cod_produto)
                else:
                    product = Product(**row_data)
                    apply_common_source_fields(product, None, "excel")
                    session.add(product)
                    created_in_batch[cod_produto] = product
                    created += 1
                    product_codes_touched.add(cod_produto)
            except Exception:
                failed += 1
                continue

        try:
            _drop_legacy_sku_unique_constraint(session)
            session.flush()
        except Exception as exc:
            session.rollback()
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Falha ao salvar importacao. Verifique duplicidades/formatos. Erro: {exc}",
            )

        _safe_log_change(
            session,
            "products",
            0,
            "import_excel",
            user.username,
            {
                "created": created,
                "updated": updated,
                "ignored": ignored,
                "failed": failed,
                "distinct_product_codes_touched": len(product_codes_touched),
            },
        )
        try:
            session.commit()
        except Exception as exc:
            session.rollback()
            logger.exception("Falha ao concluir importacao de planilha")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Falha ao concluir importacao. Erro: {exc}",
            )

        row_ops = created + updated
        total_raw = session.exec(select(func.count()).select_from(Product)).first()
        if total_raw is None:
            total_in_db = 0
        elif isinstance(total_raw, int):
            total_in_db = total_raw
        else:
            total_in_db = int(total_raw[0])

        return {
            "created": created,
            "updated": updated,
            "ignored": ignored,
            "failed": failed,
            "distinct_product_codes_touched": len(product_codes_touched),
            "rows_applied": row_ops,
            "total_products_in_db": total_in_db,
        }
    except HTTPException:
        raise
    except SQLAlchemyError as exc:
        session.rollback()
        logger.exception("Erro SQL na importacao de planilha")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Falha de banco durante importacao. Erro: {exc}",
        )
    except Exception as exc:
        session.rollback()
        logger.exception("Falha inesperada na importacao de planilha")
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Falha inesperada ao importar planilha. Erro: {exc}",
        )


def _record_history(session: Session, product_id: int, field: str, old_val, new_val, actor: str) -> None:
    session.add(ProductHistory(
        product_id=product_id,
        field_name=field,
        old_value=str(old_val) if old_val is not None else None,
        new_value=str(new_val) if new_val is not None else None,
        changed_by=actor,
    ))


@router.get("/{product_id}", response_model=ProductRead)
def get_product(
    product_id: int,
    session: Session = Depends(get_session),
    _: User = Depends(require_cadastro_access),
) -> Product:
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")
    if not (product.cod_produto or "").strip():
        product.cod_produto = (product.cod_grup_sku or str(product.id or "")).strip()
    return product


@router.put("/{product_id}", response_model=ProductRead)
def update_product(
    product_id: int,
    payload: ProductUpdate,
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> Product:
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")

    update_data = payload.model_dump(exclude_unset=True)
    if "cod_produto" in update_data and update_data["cod_produto"] is not None:
        new_cod = str(update_data["cod_produto"]).strip()
        if new_cod and new_cod != (product.cod_produto or "").strip():
            taken = session.exec(
                select(Product).where(Product.cod_produto == new_cod, Product.id != product.id),
            ).first()
            if taken:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail="Codigo do produto ja cadastrado",
                )

    for field, new_value in update_data.items():
        old_value = getattr(product, field, None)
        if str(old_value) != str(new_value):
            _record_history(session, product.id, field, old_value, new_value, user.username)
            setattr(product, field, new_value)

    from app.models.entities import utcnow
    product.updated_at = utcnow()
    session.add(product)
    session.flush()
    _safe_log_change(session, "products", product.id, "update", user.username, update_data)
    session.commit()
    session.refresh(product)
    return product


@router.patch("/{product_id}/toggle-status", response_model=ProductRead)
def toggle_product_status(
    product_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> Product:
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")

    old_status = product.status
    new_status = "inativo" if (old_status or "").lower() != "inativo" else "ativo"
    _record_history(session, product.id, "status", old_status, new_status, user.username)
    product.status = new_status

    from app.models.entities import utcnow
    product.updated_at = utcnow()
    session.add(product)
    session.flush()
    _safe_log_change(session, "products", product.id, "toggle_status", user.username, {"old": old_status, "new": new_status})
    session.commit()
    session.refresh(product)
    return product


@router.delete("/{product_id}")
def delete_product(
    product_id: int,
    session: Session = Depends(get_session),
    user: User = Depends(require_cadastro_access),
) -> dict:
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")

    _safe_log_change(
        session,
        "products",
        product.id,
        "delete",
        user.username,
        {"cod_produto": product.cod_produto, "cod_grup_sku": product.cod_grup_sku},
    )
    session.delete(product)
    session.commit()
    return {"ok": True}


@router.get("/{product_id}/history", response_model=list[ProductHistoryRead])
def get_product_history(
    product_id: int,
    session: Session = Depends(get_session),
    _: User = Depends(require_cadastro_access),
) -> list[ProductHistory]:
    product = session.get(Product, product_id)
    if not product:
        raise HTTPException(status_code=404, detail="Produto nao encontrado")

    rows = session.exec(
        select(ProductHistory)
        .where(ProductHistory.product_id == product_id)
        .order_by(ProductHistory.changed_at.desc())
    ).all()
    return list(rows)
